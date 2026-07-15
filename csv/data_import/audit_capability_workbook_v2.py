#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Audit workbook Danh mục Năng lực v2: counts, distinct masters, capability cells hợp lệ.

  python3 audit_capability_workbook_v2.py
  python3 audit_capability_workbook_v2.py --xlsx path/to/file.xlsx --json-out audit.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import openpyxl
except ImportError:
    print("Cần: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_REPO = Path(__file__).resolve().parents[3]
_DEFAULT_GLOB = list((_REPO / "data").glob("*N*ng l*c v2.xlsx"))
DEFAULT_XLSX = _DEFAULT_GLOB[0] if _DEFAULT_GLOB else _REPO / "data" / "Danh mục Năng lực v2.xlsx"

sys.path.insert(0, str(Path(__file__).resolve().parent))
import capability_header_maps as chm

_RE_CT = re.compile(r"(?i)^CT-\S")


def norm_header(h: Any) -> str:
    return chm.norm_header_cell(h)


def build_header_map(header: Tuple[Any, ...]) -> Dict[str, int]:
    pairs = [(norm_header(v), i) for i, v in enumerate(header)]

    def find(pred) -> Optional[int]:
        for nh, idx in pairs:
            if nh and pred(nh):
                return idx
        return None

    m: Dict[str, int] = {}
    for key, pred in [
        ("code", lambda nh: "mã chỉ tiêu" in nh),
        ("sm_group", lambda nh: nh == "nhóm nền mẫu"),
        ("sm", lambda nh: nh == "nền mẫu"),
        ("standard", lambda nh: "tiêu chuẩn" in nh and ("qui chuẩn" in nh or "quy chuẩn" in nh)),
        ("analysis_group", lambda nh: nh == "nhóm chỉ tiêu"),
        ("name_vi", lambda nh: nh == "tên chỉ tiêu"),
        ("name_en", lambda nh: nh == "tên tiếng anh"),
        ("short_name", lambda nh: nh == "tên viết tắt"),
        ("nd107_hcm", chm.matches_nd107_hcm),
        ("nd107_ct", chm.matches_nd107_ct),
        ("nd107_bl", chm.matches_nd107_bl),
        ("nd107_cm", chm.matches_nd107_cm),
        ("ntp_label", lambda nh: "ghi chú" in nh and "tên ntp" in nh),
        (
            "nd107_ntp",
            lambda nh: ("nđ 107" in nh or "nđ107" in nh.replace(" ", ""))
            or (
                ("chứng nhận hoạt động" in nh or "chung nhan hoat dong" in nh)
                and ("nđ107" in nh.replace(" ", "") or "nd107" in nh.replace(" ", ""))
            ),
        ),
    ]:
        idx = find(pred)
        if idx is not None:
            m[key] = idx

    iso_idxs = sorted(idx for nh, idx in pairs if "iso" in nh and "(" in nh and "a" in nh)
    if len(iso_idxs) >= 1:
        m["iso_hcm"] = iso_idxs[0]
    if len(iso_idxs) >= 2:
        m["iso_ct"] = iso_idxs[1]

    for nh, idx in pairs:
        if "cục bvtv" in nh or "cuc bvtv" in nh:
            m["cuc_bvtv"] = idx
        elif "bộ công thương" in nh or "bo cong thuong" in nh:
            m["bo_cong_thuong"] = idx
        elif "nafi" in nh:
            m["nafi"] = idx
        elif "chăn nuôi" in nh or "chan nuoi" in nh:
            m["cuc_chan_nuoi"] = idx
    return m


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def is_chua_co(s: str) -> bool:
    t = s.strip().lower()
    return not t or "chưa có" in t or "chua co" in t or t in ("na", "n/a", "-", "x", "**x**")


def is_valid_cap_cell_vietlabs(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, datetime):
        return True
    s = cell_str(v)
    if is_chua_co(s):
        return False
    return bool(re.search(r"\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}", s))


def is_valid_cap_cell_ntp(v: Any) -> bool:
    if v is None:
        return False
    s = cell_str(v).lower().replace("*", "").strip()
    if not s or is_chua_co(s):
        return False
    if s == "x":
        return True
    return bool(re.search(r"\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}", s))


def audit_sheet(ws, sheet_name: str) -> Dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"sheet": sheet_name, "data_rows": 0}
    header = tuple(rows[0])
    cmap = build_header_map(header)
    data = rows[1:]

    codes: List[str] = []
    distinct: Dict[str, Set[str]] = {
        "sm_group": set(),
        "sm": set(),
        "standard": set(),
        "analysis_group": set(),
        "equipment": set(),
        "reference_method": set(),
        "uom": set(),
        "lab_technique": set(),
        "ntp_label": set(),
    }
    cap_counts: Counter = Counter()

    eq_i = next((i for nh, i in [(norm_header(h), j) for j, h in enumerate(header)] if "thiết bị" in nh or "equipment" in nh), None)
    rm_i = next((i for nh, i in [(norm_header(h), j) for j, h in enumerate(header)] if nh == "phương pháp"), None)
    uom_i = next((i for nh, i in [(norm_header(h), j) for j, h in enumerate(header)] if nh == "đvt"), None)
    lt_i = cmap.get("lab_technique") or next(
        (i for nh, i in [(norm_header(h), j) for j, h in enumerate(header)] if nh == "kỹ thuật" or "bộ phận phụ trách" in nh),
        None,
    )

    for row in data:
        if not row or not any(v is not None and str(v).strip() for v in row):
            continue
        ci = cmap.get("code")
        if ci is None or ci >= len(row):
            continue
        code = cell_str(row[ci])
        if not code or not _RE_CT.match(code):
            continue
        codes.append(code)

        def grab(key: str, bucket: str) -> None:
            idx = cmap.get(key)
            if idx is None or idx >= len(row):
                return
            s = cell_str(row[idx])
            if s and s.upper() != "NA":
                distinct[bucket].add(s)

        grab("sm_group", "sm_group")
        grab("sm", "sm")
        grab("standard", "standard")
        grab("analysis_group", "analysis_group")
        grab("ntp_label", "ntp_label")
        if eq_i is not None and eq_i < len(row):
            s = cell_str(row[eq_i])
            if s and s.upper() != "NA":
                distinct["equipment"].add(s)
        if rm_i is not None and rm_i < len(row):
            s = cell_str(row[rm_i])
            if s:
                distinct["reference_method"].add(s)
        if uom_i is not None and uom_i < len(row):
            s = cell_str(row[uom_i])
            if s and s.upper() != "NA":
                distinct["uom"].add(s)
        if lt_i is not None and lt_i < len(row):
            s = cell_str(row[lt_i])
            if s:
                distinct["lab_technique"].add(s)

        is_ntp = sheet_name.upper().startswith("NTP")
        cap_fn = is_valid_cap_cell_ntp if is_ntp else is_valid_cap_cell_vietlabs
        for cap_key in ("nd107_hcm", "nd107_ct", "nd107_bl", "nd107_cm", "nd107_ntp"):
            idx = cmap.get(cap_key)
            if idx is not None and idx < len(row) and cap_fn(row[idx]):
                cap_counts[cap_key if cap_key != "nd107_ntp" else "nd107_hcm"] += 1
        for cap_key in ("iso_hcm", "iso_ct", "cuc_bvtv", "bo_cong_thuong", "nafi", "cuc_chan_nuoi"):
            idx = cmap.get(cap_key)
            if idx is not None and idx < len(row) and cap_fn(row[idx]):
                cap_counts[cap_key] += 1

    return {
        "sheet": sheet_name,
        "data_rows": len(codes),
        "unique_codes": len(set(codes)),
        "duplicate_code_count": len(codes) - len(set(codes)),
        "distinct_counts": {k: len(v) for k, v in distinct.items()},
        "capability_cells": dict(cap_counts),
        "sample_distinct": {k: sorted(v)[:5] for k, v in distinct.items() if v},
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--json-out", type=Path, default=None)
    args = ap.parse_args()
    path: Path = args.xlsx
    if not path.is_file():
        print(f"Không tìm thấy: {path}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    report: Dict[str, Any] = {
        "file": str(path),
        "audited_at": datetime.now(timezone.utc).isoformat(),
        "sheets": [],
    }
    all_codes: Set[str] = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        part = audit_sheet(ws, sn)
        report["sheets"].append(part)
        print(json.dumps(part, ensure_ascii=False, indent=2))
    wb.close()

    # cross-sheet unique CT
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        cmap = build_header_map(tuple(header))
        ci = cmap.get("code")
        if ci is None:
            continue
        for row in it:
            if ci < len(row):
                c = cell_str(row[ci])
                if _RE_CT.match(c):
                    all_codes.add(c)
    wb2.close()
    report["total_unique_ct_codes"] = len(all_codes)
    print(f"\nTổng mã CT unique (Vietlabs + NTP): {len(all_codes)}")

    if args.json_out:
        args.json_out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Đã ghi: {args.json_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
