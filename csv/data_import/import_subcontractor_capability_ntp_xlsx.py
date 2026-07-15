#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import nang luc NTP -> subcontractor_capability (+ designation).

Sheet NTP (ô NĐ 107 / chung cho cot dang ngay):
- Gia tri **X** (hoac *x*): co dang ky — ghi DB voi ngay het han = NULL.
- Ngay hop le: co dang ky + ngay het han.
- Chua co / rong / na: bo qua (khong xoa ban ghi cu).
"""
from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import import_analysis_item as iai
import capability_header_maps as chm
import capability_workbook_paths as cwp
try:
    import openpyxl
except ImportError:
    print("Can cai: pip install openpyxl")
    sys.exit(1)

DEFAULT_XLSX = cwp.resolve_default_capability_xlsx()
IMPORT_BRANCH_SG_ONLY = True
SITES_ND107_FULL: List[Tuple[str, str, str]] = [
    ("HCM", "nd107_hcm", "SG"), ("CT", "nd107_ct", "CT"),
    ("BL", "nd107_bl", "BL"), ("CM", "nd107_cm", "CM"),
]
DESIGNATION_CODES = {
    "iso": "ISO", "cuc_bvtv": "CUC_BVTV", "bo_cong_thuong": "BO_CONG_THUONG",
    "nafi": "NAFI", "cuc_chan_nuoi": "CUC_CHAN_NUOI",
}

def default_failure_log_path(xlsx_path: str) -> str:
    d = os.path.dirname(os.path.abspath(xlsx_path))
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(d, f"{base}_ntp_subcontractor_cap_failures_{ts}.csv")

def norm_header_cell(h: Any) -> str:
    return chm.norm_header_cell(h)

def build_column_maps(header_row: Tuple[Any, ...]) -> Tuple[Dict[str, int], Dict[str, int]]:
    pairs = [(norm_header_cell(v), i) for i, v in enumerate(header_row)]
    def find(pred):
        for nh, idx in pairs:
            if nh and pred(nh):
                return idx
        return None
    colmap: Dict[str, int] = {}
    c = find(lambda nh: "mã chỉ tiêu" in nh)
    if c is not None:
        colmap["code"] = c
    ntp_i = None
    for nh, idx in pairs:
        if "ghi chú" in nh and "tên ntp" in nh:
            ntp_i = idx
            break
        if "ghi chu" in nh and "ten ntp" in nh:
            ntp_i = idx
            break
    if ntp_i is None:
        for nh, idx in pairs:
            if "ten ntp" in nh or "tên ntp" in nh:
                ntp_i = idx
                break
    if ntp_i is not None:
        colmap["ntp_label"] = ntp_i
    cap = chm.map_nd107_columns(pairs)
    chm.map_designation_columns(pairs, cap)
    if "nd107_hcm" not in cap:
        for nh, idx in pairs:
            if ("chứng nhận hoạt động" in nh or "chung nhan hoat dong" in nh) and (
                "nđ107" in nh.replace(" ", "") or "nd107" in nh.replace(" ", "")
            ):
                cap["nd107_hcm"] = idx
                break
    if "nd107_hcm" not in cap:
        for nh, idx in pairs:
            if ("nđ 107" in nh or "nđ107" in nh.replace(" ", "") or "đ 107" in nh) and "năng lực" not in nh:
                if nh.strip() == "nđ 107" or ("ct" not in nh and "bl" not in nh and "cm" not in nh and "hcm" not in nh):
                    cap["nd107_hcm"] = idx
                    break
    return colmap, cap

def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()

def is_chua_co(s: str) -> bool:
    """Khong coi 'x' la Chua co — x la co dang ky (khong ngay)."""
    t = s.strip().lower()
    return not t or "chưa có" in t or "chua co" in t or t in ("na", "n/a", "-")


def normalize_x_cell(s: str) -> str:
    return re.sub(r"[\*]+", "", s).strip().lower()


def is_x_tick(s: str) -> bool:
    return normalize_x_cell(s) == "x"


def parse_vn_date_string(s: str) -> Optional[datetime]:
    s = s.strip()
    for sep in ("/", "-", "."):
        if sep in s:
            parts = re.split(r"[/.\-]", s)
            if len(parts) == 3:
                try:
                    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return datetime(y, m, d)
                except (ValueError, OverflowError):
                    return None
    return None


@dataclass(frozen=True)
class RegisteredCell:
    """expired_date None = co dang ky nhung khong ngay het han (ô X)."""

    expired_date: Optional[datetime]


def parse_registered_cap_cell(v: Any) -> Optional[RegisteredCell]:
    """NTP: X -> RegisteredCell(None); ngay -> RegisteredCell(dt); Chua co -> None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return RegisteredCell(v)
    s = cell_str(v)
    if not s.strip():
        return None
    if is_x_tick(s):
        return RegisteredCell(None)
    if is_chua_co(s):
        return None
    dt = parse_vn_date_string(s)
    return RegisteredCell(dt) if dt else None


def merge_iso_hcm_ct(
    rh: Optional[RegisteredCell],
    rct: Optional[RegisteredCell],
) -> Optional[RegisteredCell]:
    if rh is None and rct is None:
        return None
    dates = [
        r.expired_date
        for r in (rh, rct)
        if r is not None and r.expired_date is not None
    ]
    if dates:
        return RegisteredCell(max(dates))
    return RegisteredCell(None)


def assert_subcontractor_capability_schema(cur) -> None:
    cur.execute("""SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'subcontractor_capability'""")
    cols = {str(r[0]).lower() for r in cur.fetchall()}
    missing = {"nd_107", "nd_107_expired_date"} - cols
    if missing:
        print("Loi: bang subcontractor_capability thieu cot:", ", ".join(sorted(missing)))
        sys.exit(1)

def load_designation_ids(cur) -> Dict[str, str]:
    cur.execute("SELECT designation_code, designation_id FROM designation WHERE designation_code IS NOT NULL")
    m: Dict[str, str] = {}
    for code, did in cur.fetchall():
        if code:
            m[str(code).strip().upper()] = str(did)
    return m

def fetch_analysis_item_id(cur, analysis_item_code: str) -> Optional[str]:
    cur.execute("SELECT CAST(analysis_item_id AS nvarchar(50)) FROM analysis_item WHERE analysis_item_code = ?", analysis_item_code)
    r = cur.fetchone()
    return str(r[0]).strip() if r else None

def resolve_subcontractor_id(cur, label_raw: str) -> Optional[str]:
    label = (label_raw or "").strip()
    if not label:
        return None
    needle = label.casefold()
    cur.execute("""SELECT CAST(s.subcontractor_id AS nvarchar(50)) FROM subcontractor s
        WHERE (s.status IS NULL OR s.status = N'Active') AND s.short_name IS NOT NULL
          AND LTRIM(RTRIM(s.short_name)) <> N'' AND LOWER(LTRIM(RTRIM(s.short_name))) = ?""", needle)
    rows = cur.fetchall()
    if not rows:
        return None
    if len(rows) > 1:
        return "__ambiguous__"
    return str(rows[0][0]).strip()


def _next_auto_subcontractor_code(cur) -> str:
    cur.execute(
        """SELECT code FROM subcontractor
        WHERE code LIKE N'NTP-AUTO-%' ORDER BY code DESC"""
    )
    rows = cur.fetchall()
    max_n = 0
    for (code,) in rows:
        if not code:
            continue
        m = re.match(r"(?i)NTP-AUTO-(\d+)$", str(code).strip())
        if m:
            max_n = max(max_n, int(m.group(1)))
    return f"NTP-AUTO-{max_n + 1:04d}"


def ensure_subcontractor_id(
    cur,
    label_raw: str,
    dry_run: bool,
    created: List[Tuple[str, str, str]],
    cache: Optional[Dict[str, str]] = None,
) -> Optional[str]:
    """Tra short_name; nếu chưa có thì tạo subcontractor tối thiểu (code, short_name, name)."""
    label = (label_raw or "").strip()
    if not label:
        return None
    cache_key = label.casefold()
    if cache is not None and cache_key in cache:
        return cache[cache_key]

    sub_id = resolve_subcontractor_id(cur, label_raw)
    if sub_id == "__ambiguous__":
        return sub_id
    if sub_id:
        if cache is not None:
            cache[cache_key] = sub_id
        return sub_id

    code = _next_auto_subcontractor_code(cur)
    new_id = str(uuid.uuid4())
    if dry_run:
        created.append((code, label, new_id))
        if cache is not None:
            cache[cache_key] = new_id
        return new_id
    now = datetime.now(timezone.utc)
    cur.execute(
        """INSERT INTO subcontractor (
            subcontractor_id, code, short_name, name, status, notes, created_at
        ) VALUES (?, ?, ?, ?, N'Active', N'Tu dong tao tu import NTP', ?)""",
        (new_id, code, label, label, now),
    )
    created.append((code, label, new_id))
    if cache is not None:
        cache[cache_key] = new_id
    return new_id

def find_sc_id(cur, subcontractor_id: str, analysis_item_id: str) -> Optional[str]:
    cur.execute("SELECT subcontractor_capability_id FROM subcontractor_capability WHERE subcontractor_id = ? AND analysis_item_id = ?",
        (subcontractor_id, analysis_item_id))
    r = cur.fetchone()
    return str(r[0]) if r else None

def upsert_sc_nd107_registered(
    cur,
    subcontractor_id: str,
    analysis_item_id: str,
    nd107_expired_date: Optional[datetime],
    dry_run: bool,
) -> Tuple[str, bool]:
    """nd_107 luon True; nd107_expired_date co the NULL (ô X)."""
    now = datetime.now(timezone.utc)
    row = find_sc_id(cur, subcontractor_id, analysis_item_id)
    if row:
        sc_id = row
        if not dry_run:
            cur.execute(
                """UPDATE subcontractor_capability SET nd_107 = ?, nd_107_expired_date = ?, updated_at = ?
                WHERE subcontractor_capability_id = ?""",
                (True, nd107_expired_date, now, sc_id),
            )
        return sc_id, False
    sc_id = str(uuid.uuid4())
    if not dry_run:
        cur.execute(
            """INSERT INTO subcontractor_capability (
                subcontractor_capability_id, subcontractor_id, analysis_item_id,
                nd_107, nd_107_expired_date, status, created_at) VALUES (?, ?, ?, ?, ?, N'Active', ?)""",
            (sc_id, subcontractor_id, analysis_item_id, True, nd107_expired_date, now),
        )
    return sc_id, True

def ensure_sc_without_nd107(cur, subcontractor_id: str, analysis_item_id: str, dry_run: bool) -> Tuple[str, bool]:
    ex = find_sc_id(cur, subcontractor_id, analysis_item_id)
    if ex:
        return ex, False
    sc_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    if not dry_run:
        cur.execute("""INSERT INTO subcontractor_capability (
                subcontractor_capability_id, subcontractor_id, analysis_item_id,
                nd_107, nd_107_expired_date, status, created_at) VALUES (?, ?, ?, 0, NULL, N'Active', ?)""",
            (sc_id, subcontractor_id, analysis_item_id, now))
    return sc_id, True

def sync_sc_designation(
    cur,
    sc_id: str,
    designation_id: str,
    expired: Optional[datetime],
    dry_run: bool,
) -> None:
    """expired None = co chi dinh (ô X), ngay het han NULL."""
    cur.execute(
        """SELECT subcontractor_capability_designation_id FROM subcontractor_capability_designation
        WHERE subcontractor_capability_id = ? AND designation_id = ?""",
        (sc_id, designation_id),
    )
    ex = cur.fetchone()
    if ex:
        if not dry_run:
            cur.execute(
                "UPDATE subcontractor_capability_designation SET expired_date = ? WHERE subcontractor_capability_designation_id = ?",
                (expired, str(ex[0])),
            )
    elif not dry_run:
        did = str(uuid.uuid4())
        cur.execute(
            """INSERT INTO subcontractor_capability_designation (
            subcontractor_capability_designation_id, subcontractor_capability_id, designation_id, expired_date)
            VALUES (?, ?, ?, ?)""",
            (did, sc_id, designation_id, expired),
        )

def process(xlsx_path: str, sheet_name: str, connection, dry_run: bool, max_rows: Optional[int],
            only_codes: Optional[List[str]], failure_log_path: Optional[str], all_branches: bool) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Loi: sheet '{sheet_name}' khong co. Co: {wb.sheetnames}")
        wb.close()
        return
    ws = wb[sheet_name]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    colmap, capmap = build_column_maps(tuple(header))
    if "code" not in colmap:
        print("Loi: thieu cot Ma chi tieu")
        wb.close()
        return
    if "ntp_label" not in colmap:
        print("Loi: thieu cot Ghi chu (Ten NTP)")
        wb.close()
        return
    cur = connection.cursor()
    assert_subcontractor_capability_schema(cur)
    des_ids = load_designation_ids(cur)
    only_cf = {c.strip().casefold() for c in only_codes} if only_codes else None
    stats = {
        "rows_ok": 0,
        "skip": 0,
        "error": 0,
        "sc_insert": 0,
        "sc_update": 0,
        "scd_write": 0,
        "scd_skip_chua_co": 0,
        "subcontractor_created": 0,
    }
    created_subcontractors: List[Tuple[str, str, str]] = []
    subcontractor_cache: Dict[str, str] = {}
    failure_rows: List[Tuple[int, str, str]] = []
    def log_issue(row_idx: int, item_code: str, reason: str) -> None:
        failure_rows.append((row_idx, item_code, reason))

    def gv(key: str, row: Tuple[Any, ...]) -> str:
        i = colmap.get(key)
        return cell_str(row[i]) if i is not None and i < len(row) else ""

    def cv(key: str, row: Tuple[Any, ...]) -> Any:
        i = capmap.get(key)
        return row[i] if i is not None and i < len(row) else None

    for ridx, row in enumerate(it, start=2):
        if max_rows is not None and stats["rows_ok"] >= max_rows:
            break
        if not row:
            stats["skip"] += 1
            continue
        code = gv("code", row).strip()
        if not code:
            stats["skip"] += 1
            continue
        if only_cf is not None and code.casefold() not in only_cf:
            continue
        ntp_lbl = gv("ntp_label", row).strip()
        if not ntp_lbl:
            log_issue(ridx, code, "Cot Ghi chu (Ten NTP) trong")
            stats["error"] += 1
            continue
        sub_id = ensure_subcontractor_id(
            cur, ntp_lbl, dry_run, created_subcontractors, subcontractor_cache
        )
        if sub_id is None:
            log_issue(ridx, code, f"Cot Ghi chu (Ten NTP) trong hoac khong hop le")
            stats["error"] += 1
            continue
        if sub_id == "__ambiguous__":
            log_issue(ridx, code, f"Nhieu subcontractor khop short_name: {ntp_lbl!r}")
            stats["error"] += 1
            continue
        analysis_item_id = fetch_analysis_item_id(cur, code)
        if not analysis_item_id:
            log_issue(ridx, code, "Chua co analysis_item trong DB")
            stats["error"] += 1
            continue
        sc_id_for_des: Optional[str] = None
        try:
            if all_branches:
                nd_regs: List[RegisteredCell] = []
                for _, cap_key, _ in SITES_ND107_FULL:
                    if cap_key not in capmap:
                        continue
                    r = parse_registered_cap_cell(cv(cap_key, row))
                    if r:
                        nd_regs.append(r)
                if nd_regs:
                    dts = [r.expired_date for r in nd_regs if r.expired_date is not None]
                    nd_exp = max(dts) if dts else None
                    sid, ins = upsert_sc_nd107_registered(
                        cur, sub_id, analysis_item_id, nd_exp, dry_run
                    )
                    sc_id_for_des = sid
                    stats["sc_insert" if ins else "sc_update"] += 1
            else:
                if "nd107_hcm" in capmap:
                    nd_res = parse_registered_cap_cell(cv("nd107_hcm", row))
                    if nd_res:
                        sid, ins = upsert_sc_nd107_registered(
                            cur,
                            sub_id,
                            analysis_item_id,
                            nd_res.expired_date,
                            dry_run,
                        )
                        sc_id_for_des = sid
                        stats["sc_insert" if ins else "sc_update"] += 1

            def hcm_has_designation_write() -> bool:
                for ck in ("iso_hcm", "cuc_bvtv", "bo_cong_thuong", "nafi", "cuc_chan_nuoi"):
                    if ck not in capmap:
                        continue
                    if parse_registered_cap_cell(cv(ck, row)) is not None:
                        return True
                return False

            if sc_id_for_des is None and hcm_has_designation_write():
                sc_id_for_des, ins2 = ensure_sc_without_nd107(
                    cur, sub_id, analysis_item_id, dry_run
                )
                if ins2:
                    stats["sc_insert"] += 1

            ct_id_sc: Optional[str] = None
            if all_branches and "iso_ct" in capmap:
                if parse_registered_cap_cell(cv("iso_ct", row)) is not None:
                    ct_id_sc = find_sc_id(cur, sub_id, analysis_item_id)
                    if ct_id_sc is None:
                        ct_id_sc, ins_ct = ensure_sc_without_nd107(
                            cur, sub_id, analysis_item_id, dry_run
                        )
                        if ins_ct:
                            stats["sc_insert"] += 1

            if sc_id_for_des:
                hcm_specs = [
                    ("iso_hcm", "iso"),
                    ("cuc_bvtv", "cuc_bvtv"),
                    ("bo_cong_thuong", "bo_cong_thuong"),
                    ("nafi", "nafi"),
                    ("cuc_chan_nuoi", "cuc_chan_nuoi"),
                ]
                for ck, dk in hcm_specs:
                    if ck not in capmap:
                        continue
                    dcode = DESIGNATION_CODES[dk]
                    did = des_ids.get(dcode)
                    if ck == "iso_hcm" and all_branches and "iso_ct" in capmap:
                        reg = merge_iso_hcm_ct(
                            parse_registered_cap_cell(cv("iso_hcm", row)),
                            parse_registered_cap_cell(cv("iso_ct", row)),
                        )
                    else:
                        reg = parse_registered_cap_cell(cv(ck, row))
                    if reg is None:
                        stats["scd_skip_chua_co"] += 1
                        continue
                    if not did:
                        log_issue(ridx, code, f"Chi dinh ({ck}): thieu ma {dcode}")
                        continue
                    sync_sc_designation(
                        cur, sc_id_for_des, did, reg.expired_date, dry_run
                    )
                    stats["scd_write"] += 1

                if all_branches and "iso_ct" in capmap and "iso_hcm" not in capmap:
                    did_iso = des_ids.get("ISO")
                    reg_ct = parse_registered_cap_cell(cv("iso_ct", row))
                    if reg_ct is not None:
                        if did_iso:
                            sync_sc_designation(
                                cur, sc_id_for_des, did_iso, reg_ct.expired_date, dry_run
                            )
                            stats["scd_write"] += 1
                        else:
                            log_issue(ridx, code, "ISO CT: thieu ma ISO")

            elif hcm_has_designation_write():
                log_issue(ridx, code, "Co chi dinh HCM nhung khong tao SC")

            if (
                all_branches
                and ct_id_sc
                and "iso_ct" in capmap
                and parse_registered_cap_cell(cv("iso_ct", row)) is not None
                and not sc_id_for_des
            ):
                log_issue(ridx, code, "ISO CT nhung khong co SC")

            if not dry_run:
                connection.commit()
            stats["rows_ok"] += 1
        except Exception as e:
            stats["error"] += 1
            print(f"  Dong {ridx} {code}: {e}")
            log_issue(ridx, code, f"Exception: {e!s}")
            try:
                connection.rollback()
            except Exception:
                pass

    wb.close()
    if failure_log_path and failure_rows:
        with open(failure_log_path, "w", encoding="utf-8-sig", newline="") as lf:
            w = csv.writer(lf)
            w.writerow(["sheet_row", "analysis_item_code", "reason"])
            w.writerows(failure_rows)
    stats["subcontractor_created"] = len(created_subcontractors)
    print("\nKet qua import nang luc NTP (subcontractor_capability):")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if created_subcontractors:
        print("\nSubcontractor tu dong tao:")
        for code, label, sid in created_subcontractors:
            print(f"  {code} | short_name={label!r} | id={sid}")
    if failure_log_path:
        print(f"\nLog: {failure_log_path}" if failure_rows else "\nKhong co dong loi trong log.")

def main():
    parser = argparse.ArgumentParser(description="Import nang luc NTP")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX if os.path.isfile(DEFAULT_XLSX) else None)
    parser.add_argument("--sheet", default="NTP")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--max-rows", type=int, default=None, metavar="N")
    parser.add_argument("--only-code", action="append", dest="only_codes", metavar="CODE")
    parser.add_argument("--failure-log", default=None)
    parser.add_argument("--no-failure-log", action="store_true")
    parser.add_argument("--all-branches", action="store_true")
    args = parser.parse_args()
    args.xlsx = cwp.resolve_xlsx_arg(args.xlsx)
    if not args.xlsx or not os.path.isfile(args.xlsx):
        print("Loi: can --xlsx")
        sys.exit(1)
    if sys.platform == "win32":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    all_branches = args.all_branches or not IMPORT_BRANCH_SG_ONLY
    print("=" * 60)
    print(f"Sheet: {args.sheet} | dry_run={args.dry_run} | max_rows={args.max_rows}")
    print("=" * 60)
    fl_path = None if args.no_failure_log else (args.failure_log or default_failure_log_path(args.xlsx))
    conn = iai.pyodbc.connect(iai.CONNECTION_STRING)
    try:
        process(args.xlsx, args.sheet, conn, args.dry_run, args.max_rows, args.only_codes, fl_path, all_branches)
    finally:
        conn.close()

if __name__ == "__main__":
    main()
