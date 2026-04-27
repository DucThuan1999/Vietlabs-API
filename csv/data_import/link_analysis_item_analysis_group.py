#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Backfill relationship AnalysisItem -> AnalysisGroup in SQL Server using Capability.xlsx.

Use-case (requested):
- Data for both tables already exists in SQL Server.
- analysis_item.analysis_group_id is missing / incorrect.
- Source-of-truth for group name is Capability.xlsx (sheet Vietlabs) columns:
  - "Mã chỉ tiêu" (analysis_item_code, CT-...)
  - "Nhóm chỉ tiêu" (analysis group name)

Strategy:
- Read Capability.xlsx, build map: CT-xxxx -> analysis_group_name.
- Load analysis_group rows from DB, build normalized lookup from:
  - name_vi, name_en, analysis_group_code
  - and known legacy aliases (from CapabilityAnalysisGroupsSeedData)
- For each CT code present in DB, update analysis_item.analysis_group_id to matched group id.

Notes:
- Default behavior updates only rows with NULL/zero/invalid analysis_group_id.
- Use --dry-run to preview without updating.
"""

from __future__ import annotations

import argparse
import re
import sys
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import pyodbc
except ImportError:
    pyodbc = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


# Reuse the same connection string style as existing import scripts
CONNECTION_STRING = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=192.168.2.21,49938;"
    "Database=VietLabs;"
    "UID=sa;"
    "PWD=Sbt@2024;"
    "TrustServerCertificate=yes;"
    "Encrypt=yes;"
    "Login Timeout=60;"
)


ZERO_GUID = "00000000-0000-0000-0000-000000000000"

# Canonical (Excel) -> legacy DB names (same meaning) to reduce not-found.
# Copied from `Data/CapabilityAnalysisGroupsSeedData.cs`
LEGACY_NAME_ALIASES: Dict[str, List[str]] = {
    "BETA-AGONISTS": ["β-AGONISTS"],
    "BETA-LACTAM (PENICILLINS)": ["β-LACTAM (PENICILLINS)"],
    "FLUOROQUINOLONES": ["FLUOROQUINOLNES"],
    "NITROFURAN METABOLITES": ["NITROFURANS METABIOLIZE"],
    "PES-LC/MS/MS": ["PES-LC/MS/MS-1"],
    "HEAVY METAL-TP-ICP/MS": ["HEAVY METAL-ICP/MS"],
    "METAL-TP-ICP/MS": ["METAL/ICP-MS"],
}


def normalize_text(text: str) -> str:
    """
    Normalize for matching:
    - trim, upper
    - remove all non-word characters (including spaces/punct)
    """
    if not text:
        return ""
    s = str(text).strip().upper()
    s = re.sub(r"[^\w]", "", s)
    return s


def is_guid_like(value: object) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    try:
        uuid.UUID(s)
        return True
    except Exception:
        return False


def is_zero_guid(value: object) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() == ZERO_GUID


@dataclass(frozen=True)
class GroupRow:
    analysis_group_id: str
    analysis_group_code: Optional[str]
    name_vi: Optional[str]
    name_en: Optional[str]


def fetch_table_columns(conn, table: str, schema: str = "dbo") -> List[str]:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
        """,
        schema,
        table,
    )
    return [r[0] for r in cur.fetchall()]


def norm_header_cell(h: object) -> str:
    if h is None:
        return ""
    s = str(h).replace("\n", " ").replace('"', "").replace("'", "").strip()
    s = re.sub(r"\s+", " ", s).lower()
    return s


def find_col(header_row: Sequence[object], pred) -> Optional[int]:
    for i, v in enumerate(header_row):
        nh = norm_header_cell(v)
        if nh and pred(nh):
            return i
    return None


def get_vietlabs_columns(header_row: Sequence[object]) -> Tuple[int, int]:
    """
    Returns (code_col_idx, group_col_idx) for sheet Vietlabs.
    - code: "Mã chỉ tiêu"
    - group: "Nhóm chỉ tiêu"
    """
    code_i = find_col(header_row, lambda nh: "mã chỉ tiêu" in nh)
    group_i = find_col(header_row, lambda nh: nh == "nhóm chỉ tiêu" or "nhóm chỉ tiêu" in nh)
    if code_i is None:
        raise ValueError("Không tìm thấy cột 'Mã chỉ tiêu' trong header.")
    if group_i is None:
        raise ValueError("Không tìm thấy cột 'Nhóm chỉ tiêu' trong header.")
    return code_i, group_i


def is_valid_ct_code(code: str) -> bool:
    if not code:
        return False
    s = code.strip()
    return bool(re.match(r"(?i)^CT-\S", s))


def read_capability_xlsx_ct_to_group(
    xlsx_path: str,
    sheet_name: str,
) -> Dict[str, str]:
    """
    Parse Capability.xlsx and returns mapping: CT-xxxx -> group_name (as in Excel).
    First non-empty group wins per code.
    """
    if openpyxl is None:
        raise RuntimeError("Thiếu thư viện openpyxl. Cài bằng: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không thấy sheet '{sheet_name}'. Có: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    code_i, group_i = get_vietlabs_columns(header)

    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_i]).strip() if row[code_i] is not None else ""
        if not is_valid_ct_code(code):
            continue
        group = str(row[group_i]).strip() if row[group_i] is not None else ""
        if not group:
            continue
        out.setdefault(code, group)
    return out


def load_groups(conn) -> Tuple[List[GroupRow], Dict[str, str]]:
    """
    Returns:
    - list of groups
    - normalized lookup: normalized_key -> analysis_group_id
    """
    cur = conn.cursor()
    cur.execute(
        """
        SELECT analysis_group_id, analysis_group_code, name_vi, name_en
        FROM analysis_group
        WHERE status = 'Active' OR status IS NULL
        """
    )

    groups: List[GroupRow] = []
    lookup: Dict[str, str] = {}

    for gid, code, nvi, nen in cur.fetchall():
        row = GroupRow(
            analysis_group_id=str(gid),
            analysis_group_code=str(code).strip() if code is not None else None,
            name_vi=str(nvi).strip() if nvi is not None else None,
            name_en=str(nen).strip() if nen is not None else None,
        )
        groups.append(row)

        for key in (row.name_vi, row.name_en, row.analysis_group_code):
            nk = normalize_text(key or "")
            if nk:
                lookup.setdefault(nk, row.analysis_group_id)

    return groups, lookup


def build_group_key_to_ids(groups: Sequence[GroupRow]) -> Dict[str, List[str]]:
    # Use list for stable output, but dedupe IDs per key.
    key_to_ids: Dict[str, List[str]] = {}
    for g in groups:
        for k in (g.name_vi, g.name_en, g.analysis_group_code):
            nk = normalize_text(k or "")
            if nk:
                ids = key_to_ids.setdefault(nk, [])
                if g.analysis_group_id not in ids:
                    ids.append(g.analysis_group_id)
    return key_to_ids


def apply_legacy_aliases_to_key_map(key_to_ids: Dict[str, List[str]]) -> None:
    """
    If canonical name exists in DB, add keys for its legacy names to point to same id.
    """
    for canonical, legacy_list in LEGACY_NAME_ALIASES.items():
        nk_canon = normalize_text(canonical)
        ids = key_to_ids.get(nk_canon, [])
        if len(ids) != 1:
            continue
        gid = ids[0]
        for legacy in legacy_list:
            nk_legacy = normalize_text(legacy)
            if nk_legacy:
                cur = key_to_ids.setdefault(nk_legacy, [])
                if gid not in cur:
                    cur.append(gid)


def pick_preferred_group_id(groups: Sequence[GroupRow], ids: Sequence[str]) -> Optional[str]:
    """
    Deterministically pick one group id when DB has duplicate groups for same normalized key.
    Preference order:
    - Has code (not null/empty) AND smallest numeric suffix of code (e.g. NCT-0034 < NCT-0035)
    - Otherwise, first by code then by id
    """
    if not ids:
        return None
    if len(ids) == 1:
        return ids[0]

    gid_to_group: Dict[str, GroupRow] = {g.analysis_group_id: g for g in groups}

    def code_rank(code: Optional[str]) -> Tuple[int, int, str]:
        if not code:
            return (1, 10**9, "")
        s = str(code).strip().upper()
        m = re.search(r"(\d+)$", s)
        if m:
            return (0, int(m.group(1)), s)
        return (0, 10**8, s)

    ranked: List[Tuple[Tuple[int, int, str], str]] = []
    for gid in ids:
        g = gid_to_group.get(gid)
        ranked.append((code_rank(g.analysis_group_code if g else None), gid))
    ranked.sort(key=lambda x: (x[0], x[1]))
    return ranked[0][1]


def fetch_items_by_codes(
    conn,
    codes: Sequence[str],
) -> Dict[str, Tuple[str, Optional[str]]]:
    """
    Returns mapping: analysis_item_code -> (analysis_item_id, current_analysis_group_id_str)
    """
    out: Dict[str, Tuple[str, Optional[str]]] = {}
    if not codes:
        return out

    cur = conn.cursor()

    # Chunk to avoid SQL Server parameter limits
    CHUNK = 500
    for i in range(0, len(codes), CHUNK):
        chunk = list(codes[i : i + CHUNK])
        placeholders = ",".join(["?"] * len(chunk))
        sql = f"""
            SELECT
                analysis_item_code,
                CONVERT(varchar(36), analysis_item_id) AS analysis_item_id,
                CONVERT(varchar(36), analysis_group_id) AS analysis_group_id
            FROM analysis_item
            WHERE analysis_item_code IN ({placeholders})
        """
        cur.execute(sql, *chunk)
        for code, item_id, gid in cur.fetchall():
            out[str(code).strip()] = (str(item_id), str(gid).strip() if gid is not None else None)
    return out


def main() -> int:
    if pyodbc is None:
        print("Thiếu thư viện pyodbc. Cài bằng: pip install pyodbc", file=sys.stderr)
        return 2

    if openpyxl is None:
        print("Thiếu thư viện openpyxl. Cài bằng: pip install openpyxl", file=sys.stderr)
        return 2

    ap = argparse.ArgumentParser(
        description="Backfill analysis_item.analysis_group_id using Capability.xlsx (CT code -> group name)."
    )
    ap.add_argument("--dry-run", action="store_true", help="Chỉ in kết quả, không UPDATE DB.")
    ap.add_argument(
        "--xlsx",
        # Repo layout here is: <repo>/Vietlabs-API/csv/data_import/<this_file>
        # Capability.xlsx is at: <repo>/data/Capability.xlsx
        default=str(Path(__file__).resolve().parents[3] / "data" / "Capability.xlsx"),
        help="Đường dẫn Capability.xlsx (mặc định: <repo>/data/Capability.xlsx).",
    )
    ap.add_argument(
        "--sheet",
        default="Vietlabs",
        help="Tên sheet trong Capability.xlsx (mặc định: Vietlabs).",
    )
    ap.add_argument(
        "--all",
        action="store_true",
        help="Process tất cả analysis_item (kể cả đã có analysis_group_id). Mặc định chỉ xử lý rows thiếu/invalid.",
    )
    ap.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Giới hạn số record xử lý (0 = không giới hạn).",
    )
    ap.add_argument(
        "--commit-every",
        type=int,
        default=500,
        help="Commit theo batch N updates để tránh transaction quá lớn.",
    )
    args = ap.parse_args()

    print("=" * 70)
    print("Link AnalysisItem -> AnalysisGroup")
    print("=" * 70)
    print(f"Dry-run: {args.dry_run}")
    print(f"XLSX: {args.xlsx}")
    print(f"Sheet: {args.sheet}")

    conn = pyodbc.connect(CONNECTION_STRING)
    conn.autocommit = False

    try:
        xlsx_path = args.xlsx
        if not Path(xlsx_path).exists():
            print(f"Không tìm thấy file XLSX: {xlsx_path}", file=sys.stderr)
            return 1

        ct_to_group = read_capability_xlsx_ct_to_group(xlsx_path, args.sheet)
        if not ct_to_group:
            print("Không đọc được mapping CT -> Nhóm chỉ tiêu từ XLSX.", file=sys.stderr)
            return 1
        print(f"Loaded XLSX mappings: {len(ct_to_group)} CT codes with group.")

        groups, lookup = load_groups(conn)
        key_to_ids = build_group_key_to_ids(groups)
        apply_legacy_aliases_to_key_map(key_to_ids)
        print(f"Loaded analysis_group: {len(groups)} rows, normalized keys: {len(key_to_ids)}")

        # Apply limit to CT codes if requested
        codes = list(ct_to_group.keys())
        if args.limit and args.limit > 0:
            codes = codes[: int(args.limit)]

        item_by_code = fetch_items_by_codes(conn, codes)
        print(f"Matched analysis_item rows by CT code: {len(item_by_code)}")

        only_missing_or_invalid = not args.all

        updated = 0
        skipped = 0
        not_found_group = 0
        ambiguous = 0
        missing_item = 0
        invalid_group_text = 0

        cur = conn.cursor()
        commit_every = max(1, int(args.commit_every))
        start_ts = datetime.now(timezone.utc)

        for code in codes:
            group_text = (ct_to_group.get(code) or "").strip()
            if not group_text:
                invalid_group_text += 1
                continue

            item = item_by_code.get(code)
            if not item:
                missing_item += 1
                continue

            item_id, current_gid = item

            # Skip rows that already have a non-zero FK unless --all
            if only_missing_or_invalid:
                if current_gid and is_guid_like(current_gid) and not is_zero_guid(current_gid):
                    # FK exists (assume valid); skip
                    skipped += 1
                    continue

            nk = normalize_text(group_text)
            candidate_ids = key_to_ids.get(nk, [])
            if len(candidate_ids) == 0:
                not_found_group += 1
                continue
            target_gid = pick_preferred_group_id(groups, candidate_ids)
            if not target_gid:
                not_found_group += 1
                continue
            if len(candidate_ids) > 1:
                # Duplicate groups in DB for same key; we still proceed with deterministic pick.
                ambiguous += 1

            if current_gid and is_guid_like(current_gid) and str(current_gid).lower() == str(target_gid).lower():
                skipped += 1
                continue

            if args.dry_run:
                updated += 1
            else:
                cur.execute(
                    """
                    UPDATE analysis_item
                    SET analysis_group_id = ?, updated_at = ?
                    WHERE analysis_item_id = ?
                    """,
                    target_gid,
                    datetime.now(timezone.utc),
                    item_id,
                )
                updated += 1
                if updated % commit_every == 0:
                    conn.commit()
                    print(f"Committed {updated} updates...")

        if not args.dry_run:
            conn.commit()

        elapsed_s = (datetime.now(timezone.utc) - start_ts).total_seconds()
        print("\nKết quả:")
        print(f"  - Updated (or would update): {updated}")
        print(f"  - Skipped (already linked): {skipped}")
        print(f"  - Missing analysis_item by CT code: {missing_item}")
        print(f"  - Not found (no matching group in DB): {not_found_group}")
        print(f"  - Ambiguous (multiple groups share same normalized key): {ambiguous}")
        print(f"  - Invalid/empty group text from XLSX: {invalid_group_text}")
        print(f"  - Elapsed: {elapsed_s:.1f}s")

        if args.dry_run:
            print("\nGợi ý: chạy lại bỏ --dry-run để UPDATE DB.")

        return 0
    finally:
        try:
            conn.close()
        except Exception:
            pass


if __name__ == "__main__":
    raise SystemExit(main())

