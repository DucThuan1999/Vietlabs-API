#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Đồng bộ đơn giá chuẩn (analysis_item.unit_price) và giá nhóm chuẩn
(analysis_group.whole_group_standard_price) từ workbook Capability / Danh mục Năng lực.

Dùng khi import full catalog đã chạy nhưng cột giá không được map (vd. V3 dùng
"Đơn giá chuẩn" thay vì "Đơn giá chuẩn_new").

  python3 sync_unit_price_from_capability_xlsx.py --dry-run
  python3 sync_unit_price_from_capability_xlsx.py --xlsx "../../data/Danh mục Năng lực v3.xlsx"
  python3 sync_unit_price_from_capability_xlsx.py --xlsx "../../data/Danh mục Năng lực v3.xlsx" --sheet Vietlabs
"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import capability_workbook_paths as cwp
import import_analysis_item as iai
import import_analysis_items_capability_vietlabs_xlsx as cap_imp

try:
    import openpyxl
    import pyodbc
except ImportError:
    print("Cần: pip install openpyxl pyodbc", file=sys.stderr)
    sys.exit(1)

_REPO = Path(__file__).resolve().parents[3]


def resolve_v3_default() -> Optional[str]:
    data = _REPO / "data"
    if not data.is_dir():
        return None
    for f in sorted(data.iterdir()):
        if f.suffix.lower() == ".xlsx" and "v3" in f.name.lower():
            return str(f)
    return None


def load_excel_prices(
    xlsx_path: str,
    sheet_name: str,
) -> Tuple[Dict[str, float], Dict[str, List[float]], Dict[str, int]]:
    """
    Returns:
      code -> unit_price
      analysis_group_name -> list of group prices seen
      stats dict with counts
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Không thấy sheet {sheet_name!r}. Có: {', '.join(wb.sheetnames)}")

    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    colmap = cap_imp.build_column_map(header_row)

    if "unit_price" not in colmap:
        wb.close()
        raise RuntimeError(
            "Không map được cột đơn giá chuẩn. "
            f"Headers giá: {[h for h in header_row if h and 'giá' in str(h).lower()]}"
        )

    code_i = colmap["code"]
    price_i = colmap["unit_price"]
    ag_i = colmap.get("analysis_group")
    gp_i = colmap.get("group_price")

    item_prices: Dict[str, float] = {}
    group_prices: Dict[str, List[float]] = defaultdict(list)
    stats = {"rows": 0, "skipped_invalid_code": 0, "zero_price": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["rows"] += 1
        code = cap_imp.cell_str(row[code_i] if code_i < len(row) else None)
        if not cap_imp.is_valid_ct_code(code):
            stats["skipped_invalid_code"] += 1
            continue

        up_raw = row[price_i] if price_i < len(row) else None
        up_str = cap_imp.cell_str(up_raw)
        price = iai.parse_price(up_str) if up_str else 0.0
        if price is None:
            price = 0.0
        price = round(float(price), 4)
        if price == 0.0:
            stats["zero_price"] += 1
        item_prices[code.strip()] = price

        if ag_i is not None and gp_i is not None:
            ag_name = cap_imp.cell_str(row[ag_i] if ag_i < len(row) else None)
            gp_raw = row[gp_i] if gp_i < len(row) else None
            gp = iai.parse_price(cap_imp.cell_str(gp_raw)) if gp_raw is not None else None
            if ag_name and gp is not None and gp > 0:
                group_prices[ag_name.strip()].append(float(gp))

    wb.close()
    return item_prices, dict(group_prices), stats


def mode_price(values: List[float]) -> Optional[float]:
    if not values:
        return None
    from collections import Counter

    c = Counter(round(v, 4) for v in values)
    return float(c.most_common(1)[0][0])


def sync_prices(
    xlsx_path: str,
    *,
    sheet: str = "Vietlabs",
    dry_run: bool = False,
    connection_string: Optional[str] = None,
) -> Dict[str, Any]:
    item_prices, group_prices, excel_stats = load_excel_prices(xlsx_path, sheet)

    conn_str = (connection_string or iai.CONNECTION_STRING).strip()
    if "Login Timeout" not in conn_str:
        conn_str += "Login Timeout=60;"
    if "TrustServerCertificate" not in conn_str:
        conn_str += "TrustServerCertificate=yes;"
    conn = pyodbc.connect(conn_str, timeout=60)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT analysis_item_id, analysis_item_code, unit_price FROM analysis_item WHERE analysis_item_code IS NOT NULL"
        )
        db_rows = {str(r[1]).strip(): (str(r[0]), r[2]) for r in cur.fetchall()}

        updated_items = 0
        unchanged_items = 0
        missing_in_db: List[str] = []
        sample_updates: List[Dict[str, Any]] = []

        now = datetime.now(timezone.utc)
        for code, excel_price in sorted(item_prices.items()):
            if code not in db_rows:
                missing_in_db.append(code)
                continue
            aid, db_price = db_rows[code]
            db_val = round(float(db_price), 4) if db_price is not None else 0.0
            if db_val == excel_price:
                unchanged_items += 1
                continue
            if not dry_run:
                cur.execute(
                    "UPDATE analysis_item SET unit_price = ?, updated_at = ? WHERE analysis_item_id = ?",
                    excel_price,
                    now,
                    aid,
                )
            updated_items += 1
            if len(sample_updates) < 10:
                sample_updates.append({"code": code, "old": db_val, "new": excel_price})

        # Group prices
        cur.execute("SELECT analysis_group_id, name_vi, whole_group_standard_price FROM analysis_group")
        ag_by_name = {str(r[1]).strip(): (str(r[0]), r[2]) for r in cur.fetchall()}

        updated_groups = 0
        unchanged_groups = 0
        for ag_name, prices in sorted(group_prices.items()):
            target = mode_price(prices)
            if target is None:
                continue
            if ag_name not in ag_by_name:
                continue
            gid, db_gp = ag_by_name[ag_name]
            db_val = round(float(db_gp), 4) if db_gp is not None else None
            if db_val == round(target, 4):
                unchanged_groups += 1
                continue
            if not dry_run:
                cur.execute(
                    "UPDATE analysis_group SET whole_group_standard_price = ? WHERE analysis_group_id = ?",
                    target,
                    gid,
                )
            updated_groups += 1

        if not dry_run:
            conn.commit()

        return {
            "xlsx": xlsx_path,
            "sheet": sheet,
            "dry_run": dry_run,
            "excel": excel_stats,
            "excel_items_with_price": len(item_prices),
            "db_items": len(db_rows),
            "item_price": {
                "updated": updated_items,
                "unchanged": unchanged_items,
                "missing_in_db": len(missing_in_db),
                "missing_codes_sample": missing_in_db[:20],
            },
            "group_price": {
                "updated": updated_groups,
                "unchanged": unchanged_groups,
                "excel_groups": len(group_prices),
            },
            "sample_updates": sample_updates,
        }
    finally:
        conn.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Đồng bộ đơn giá chuẩn từ workbook Capability")
    parser.add_argument("--xlsx", default=None, help="Workbook (mặc định: v3 nếu có, không thì v2)")
    parser.add_argument("--sheet", default="Vietlabs")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--connection-string", default=None)
    parser.add_argument("--json-out", default=None)
    args = parser.parse_args()

    xlsx = args.xlsx
    if not xlsx or not os.path.isfile(xlsx):
        xlsx = resolve_v3_default() or cwp.resolve_default_capability_xlsx()
    if not xlsx or not os.path.isfile(xlsx):
        print("Lỗi: không tìm thấy workbook. Dùng --xlsx.", file=sys.stderr)
        return 1

    try:
        report = sync_prices(
            xlsx,
            sheet=args.sheet,
            dry_run=args.dry_run,
            connection_string=args.connection_string,
        )
    except Exception as e:
        print(f"Lỗi: {e}", file=sys.stderr)
        return 1

    mode = "DRY-RUN" if report["dry_run"] else "EXECUTE"
    print("=" * 60)
    print(f"SYNC GIÁ CHUẨN — {mode}")
    print("=" * 60)
    print(f"Workbook: {report['xlsx']}")
    print(f"Sheet: {report['sheet']}")
    ip = report["item_price"]
    gp = report["group_price"]
    print(f"\nĐơn giá chỉ tiêu (analysis_item.unit_price):")
    print(f"  Excel có: {report['excel_items_with_price']} mã")
    print(f"  DB có: {report['db_items']} mã")
    print(f"  Cập nhật: {ip['updated']}")
    print(f"  Đã khớp (bỏ qua): {ip['unchanged']}")
    print(f"  Thiếu trong DB: {ip['missing_in_db']}")
    if ip["missing_codes_sample"]:
        print(f"  Mẫu thiếu: {', '.join(ip['missing_codes_sample'][:10])}")
    print(f"\nGiá nhóm chuẩn (analysis_group.whole_group_standard_price):")
    print(f"  Cập nhật: {gp['updated']}")
    print(f"  Đã khớp: {gp['unchanged']}")
    if report["sample_updates"]:
        print("\nMẫu cập nhật:")
        for s in report["sample_updates"]:
            print(f"  {s['code']}: {s['old']} -> {s['new']}")

    if args.json_out:
        import json

        os.makedirs(os.path.dirname(os.path.abspath(args.json_out)) or ".", exist_ok=True)
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"\nĐã ghi: {args.json_out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
