#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cập nhật 6 cột location cho bảng client từ workbook Excel.

Đọc sheet Customers, match theo internal_code (cột "Mã khách hàng nội bộ").
Lấy từ Excel:
  - country, province, ward: giá trị text (Quốc gia, Tỉnh/Thành phố, Xã/Phường)
  - country_id, province_id, ward_id: uniqueidentifier
Mặc định dry-run; chỉ commit khi truyền --apply.

Usage:
  python update_client_location_ids_from_xlsx.py --xlsx ../../../data/Danh_sach_khach_hang_20260608.xlsx
  python update_client_location_ids_from_xlsx.py --xlsx ../../../data/Danh_sach_khach_hang_20260608.xlsx --apply

Optional env: VIETLABS_SQL_CONNECTION (ODBC connection string)
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple
from uuid import UUID

try:
    import openpyxl
except ImportError:
    print("Cần cài: pip install openpyxl")
    sys.exit(1)

try:
    import pyodbc
except ImportError:
    print("Cần cài: pip install pyodbc")
    sys.exit(1)


DEFAULT_XLSX = os.path.normpath(
    os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "data", "Danh_sach_khach_hang_20260608.xlsx"
    )
)

DEFAULT_CONNECTION = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=192.168.2.21,49938;"
    "Database=VietLabs;"
    "UID=sa;"
    "PWD=Sbt@2024;"
    "TrustServerCertificate=yes;"
    "Encrypt=yes;"
    "Login Timeout=120;"
)

SHEET_NAME = "Customers"
COL_INTERNAL_CODE = "Mã khách hàng nội bộ"
COL_COUNTRY = "Quốc gia"
COL_PROVINCE = "Tỉnh/Thành phố"
COL_WARD = "Xã/Phường"
COL_COUNTRY_ID = "country_id"
COL_PROVINCE_ID = "province_id"
COL_WARD_ID = "ward_id"
COL_COMPANY_NAME = "Tên công ty"

TARGET_TEXT_COLUMNS = ("country", "province", "ward")
TARGET_ID_COLUMNS = ("country_id", "province_id", "ward_id")
TARGET_COLUMNS = TARGET_TEXT_COLUMNS + TARGET_ID_COLUMNS


def connection_string() -> str:
    return os.environ.get("VIETLABS_SQL_CONNECTION", DEFAULT_CONNECTION)


def connect_sql_server(conn_s: str):
    last_exc: Optional[Exception] = None
    candidates = [conn_s]
    patched = re.sub(
        r"Encrypt\s*=\s*yes\b", "Encrypt=optional", conn_s, count=1, flags=re.I
    )
    if patched != conn_s:
        candidates.append(patched)
    for cs in candidates:
        try:
            return pyodbc.connect(cs, timeout=120)
        except pyodbc.Error as e:
            last_exc = e
    assert last_exc is not None
    raise last_exc


def strip_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def parse_uuid(value: Any) -> Optional[str]:
    raw = strip_cell(value)
    if not raw:
        return None
    try:
        return str(UUID(raw)).upper()
    except (ValueError, AttributeError):
        return None


def table_column_types(cursor, table: str) -> Dict[str, str]:
    cursor.execute(
        """
        SELECT LOWER(c.COLUMN_NAME), LOWER(c.DATA_TYPE)
        FROM INFORMATION_SCHEMA.COLUMNS c
        INNER JOIN INFORMATION_SCHEMA.TABLES t
          ON c.TABLE_SCHEMA = t.TABLE_SCHEMA AND c.TABLE_NAME = t.TABLE_NAME
        WHERE t.TABLE_SCHEMA = N'dbo'
          AND t.TABLE_NAME = ?
          AND t.TABLE_TYPE = N'BASE TABLE'
        """,
        (table,),
    )
    return {str(r[0]): str(r[1]) for r in cursor.fetchall()}


def ensure_client_location_columns(cursor) -> None:
    cols = table_column_types(cursor, "client")
    missing = [c for c in TARGET_COLUMNS if c not in cols]
    if missing:
        raise RuntimeError(
            "Bảng dbo.client chưa có các cột: "
            + ", ".join(missing)
            + ". Chạy Scripts/AddClientLocationIdColumns.sql trước."
        )
    text_types = {"nvarchar", "varchar", "nchar", "char"}
    for col in TARGET_TEXT_COLUMNS:
        if cols[col] not in text_types:
            raise RuntimeError(
                f"Cột dbo.client.{col} phải là text, thực tế: {cols[col]}"
            )
    for col in TARGET_ID_COLUMNS:
        if cols[col] != "uniqueidentifier":
            raise RuntimeError(
                f"Cột dbo.client.{col} phải là uniqueidentifier, thực tế: {cols[col]}"
            )


def normalize_text_value(value: Optional[str]) -> Optional[str]:
    """Chuẩn hóa giá trị text country/province/ward để so sánh."""
    if value is None:
        return None
    raw = strip_cell(value)
    return raw.casefold() if raw else None


def resolve_location_text(
    excel_text: Any,
    location_id: Optional[str],
    label_map: Dict[str, str],
) -> Optional[str]:
    """Ưu tiên text từ Excel; nếu trống thì fallback tên từ danh mục theo ID."""
    text = strip_cell(excel_text)
    if text:
        return text
    if location_id and location_id in label_map:
        return label_map[location_id]
    return None


def normalize_uuid_value(value: Optional[str]) -> Optional[str]:
    """Chuẩn hóa UUID từ cột uniqueidentifier hoặc text."""
    if value is None:
        return None
    raw = strip_cell(value)
    if not raw:
        return None
    parsed = parse_uuid(raw)
    return parsed if parsed else raw.upper()


def load_lookup_maps(cursor) -> Tuple[Set[str], Dict[str, str], Dict[str, Tuple[str, str]]]:
    cursor.execute("SELECT UPPER(CAST(country_id AS NVARCHAR(36))) FROM dbo.country")
    countries = {str(r[0]) for r in cursor.fetchall()}

    cursor.execute(
        "SELECT UPPER(CAST(province_id AS NVARCHAR(36))), "
        "UPPER(CAST(country_id AS NVARCHAR(36))) FROM dbo.province"
    )
    provinces: Dict[str, str] = {}
    for province_id, country_id in cursor.fetchall():
        provinces[str(province_id)] = str(country_id)

    cursor.execute(
        "SELECT UPPER(CAST(ward_id AS NVARCHAR(36))), "
        "UPPER(CAST(province_id AS NVARCHAR(36))), "
        "UPPER(CAST(country_id AS NVARCHAR(36))) FROM dbo.ward"
    )
    wards: Dict[str, Tuple[str, str]] = {}
    for ward_id, province_id, country_id in cursor.fetchall():
        wards[str(ward_id)] = (str(province_id), str(country_id))

    return countries, provinces, wards


def load_label_maps(
    cursor,
) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
    cursor.execute(
        "SELECT UPPER(CAST(country_id AS NVARCHAR(36))), full_name_vi "
        "FROM dbo.country"
    )
    country_labels = {
        str(country_id): strip_cell(label) for country_id, label in cursor.fetchall()
    }

    cursor.execute(
        "SELECT UPPER(CAST(province_id AS NVARCHAR(36))), "
        "COALESCE(NULLIF(LTRIM(RTRIM(full_name)), N''), name) "
        "FROM dbo.province"
    )
    province_labels = {
        str(province_id): strip_cell(label) for province_id, label in cursor.fetchall()
    }

    cursor.execute(
        "SELECT UPPER(CAST(ward_id AS NVARCHAR(36))), name FROM dbo.ward"
    )
    ward_labels = {
        str(ward_id): strip_cell(label) for ward_id, label in cursor.fetchall()
    }

    return country_labels, province_labels, ward_labels


ClientState = Tuple[
    str,
    Optional[str],
    Optional[str],
    Optional[str],
    Optional[str],
    Optional[str],
    Optional[str],
]


def load_clients_by_internal_code(cursor) -> Dict[str, ClientState]:
    cursor.execute(
        "SELECT UPPER(CAST(client_id AS NVARCHAR(36))), internal_code, "
        "country, province, ward, "
        "UPPER(CAST(country_id AS NVARCHAR(36))), "
        "UPPER(CAST(province_id AS NVARCHAR(36))), "
        "UPPER(CAST(ward_id AS NVARCHAR(36))) "
        "FROM dbo.client"
    )
    out: Dict[str, ClientState] = {}
    for (
        client_id,
        internal_code,
        country,
        province,
        ward,
        country_id,
        province_id,
        ward_id,
    ) in cursor.fetchall():
        code = strip_cell(internal_code)
        if not code:
            continue
        out[code] = (
            str(client_id),
            strip_cell(country) or None,
            strip_cell(province) or None,
            strip_cell(ward) or None,
            str(country_id) if country_id else None,
            str(province_id) if province_id else None,
            str(ward_id) if ward_id else None,
        )
    return out


@dataclass
class ExcelRow:
    row_no: int
    internal_code: str
    company_name: str
    country: str
    province: str
    ward: Optional[str]
    country_id: Optional[str]
    province_id: Optional[str]
    ward_id: Optional[str]


@dataclass
class RowResult:
    row_no: int
    internal_code: str
    company_name: str
    status: str
    reason: str = ""
    country_id: str = ""
    province_id: str = ""
    ward_id: str = ""


@dataclass
class RunSummary:
    total_rows: int = 0
    would_update: int = 0
    skipped: int = 0
    unchanged: int = 0
    applied: int = 0
    by_reason: Counter = field(default_factory=Counter)


def read_excel_rows(
    xlsx_path: str,
    country_labels: Dict[str, str],
    province_labels: Dict[str, str],
    ward_labels: Dict[str, str],
) -> Tuple[List[ExcelRow], List[RowResult]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Không tìm thấy sheet '{SHEET_NAME}' trong {xlsx_path}")

    ws = wb[SHEET_NAME]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' không có header")

    headers = [strip_cell(v) for v in header_row]
    header_index = {name: idx for idx, name in enumerate(headers)}

    required_headers = [
        COL_INTERNAL_CODE,
        COL_COUNTRY_ID,
        COL_PROVINCE_ID,
        COL_WARD_ID,
    ]
    missing_headers = [h for h in required_headers if h not in header_index]
    if missing_headers:
        raise RuntimeError(f"Thiếu cột bắt buộc trong Excel: {', '.join(missing_headers)}")

    rows: List[ExcelRow] = []
    issues: List[RowResult] = []
    seen_codes: Dict[str, int] = {}

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        internal_code = strip_cell(row[header_index[COL_INTERNAL_CODE]])
        company_name = strip_cell(
            row[header_index[COL_COMPANY_NAME]] if COL_COMPANY_NAME in header_index else ""
        )

        if not internal_code:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code="",
                    company_name=company_name,
                    status="skip",
                    reason="missing_internal_code",
                )
            )
            continue

        if internal_code in seen_codes:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason=f"duplicate_internal_code_first_at_row_{seen_codes[internal_code]}",
                )
            )
            continue
        seen_codes[internal_code] = row_no

        country_id = parse_uuid(row[header_index[COL_COUNTRY_ID]])
        province_id = parse_uuid(row[header_index[COL_PROVINCE_ID]])
        ward_id = parse_uuid(row[header_index[COL_WARD_ID]])

        if strip_cell(row[header_index[COL_COUNTRY_ID]]) and country_id is None:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="invalid_country_id_uuid",
                )
            )
            continue
        if strip_cell(row[header_index[COL_PROVINCE_ID]]) and province_id is None:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="invalid_province_id_uuid",
                )
            )
            continue
        if strip_cell(row[header_index[COL_WARD_ID]]) and ward_id is None:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="invalid_ward_id_uuid",
                )
            )
            continue

        if not country_id or not province_id:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="missing_country_or_province_id",
                )
            )
            continue

        country_text = resolve_location_text(
            row[header_index[COL_COUNTRY]] if COL_COUNTRY in header_index else "",
            country_id,
            country_labels,
        )
        province_text = resolve_location_text(
            row[header_index[COL_PROVINCE]] if COL_PROVINCE in header_index else "",
            province_id,
            province_labels,
        )
        ward_text = resolve_location_text(
            row[header_index[COL_WARD]] if COL_WARD in header_index else "",
            ward_id,
            ward_labels,
        ) if ward_id else None

        if not country_text:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="missing_country_text",
                    country_id=country_id,
                    province_id=province_id,
                    ward_id=ward_id or "",
                )
            )
            continue

        if not province_text:
            issues.append(
                RowResult(
                    row_no=row_no,
                    internal_code=internal_code,
                    company_name=company_name,
                    status="skip",
                    reason="missing_province_text",
                    country_id=country_id,
                    province_id=province_id,
                    ward_id=ward_id or "",
                )
            )
            continue

        rows.append(
            ExcelRow(
                row_no=row_no,
                internal_code=internal_code,
                company_name=company_name,
                country=country_text,
                province=province_text,
                ward=ward_text,
                country_id=country_id,
                province_id=province_id,
                ward_id=ward_id,
            )
        )

    wb.close()
    return rows, issues


def validate_and_plan_updates(
    excel_rows: List[ExcelRow],
    clients: Dict[str, ClientState],
    countries: Set[str],
    provinces: Dict[str, str],
    wards: Dict[str, Tuple[str, str]],
) -> Tuple[List[Tuple[str, str, str, Optional[str], str, str, Optional[str], ExcelRow]], List[RowResult], RunSummary]:
    updates: List[Tuple[str, str, str, Optional[str], str, str, Optional[str], ExcelRow]] = []
    issues: List[RowResult] = []
    summary = RunSummary(total_rows=len(excel_rows))

    for row in excel_rows:
        if row.internal_code not in clients:
            issues.append(
                RowResult(
                    row_no=row.row_no,
                    internal_code=row.internal_code,
                    company_name=row.company_name,
                    status="skip",
                    reason="client_not_found",
                    country_id=row.country_id or "",
                    province_id=row.province_id or "",
                    ward_id=row.ward_id or "",
                )
            )
            summary.skipped += 1
            summary.by_reason["client_not_found"] += 1
            continue

        if row.country_id not in countries:
            issues.append(
                RowResult(
                    row_no=row.row_no,
                    internal_code=row.internal_code,
                    company_name=row.company_name,
                    status="skip",
                    reason="country_not_found",
                    country_id=row.country_id,
                    province_id=row.province_id or "",
                    ward_id=row.ward_id or "",
                )
            )
            summary.skipped += 1
            summary.by_reason["country_not_found"] += 1
            continue

        if row.province_id not in provinces:
            issues.append(
                RowResult(
                    row_no=row.row_no,
                    internal_code=row.internal_code,
                    company_name=row.company_name,
                    status="skip",
                    reason="province_not_found",
                    country_id=row.country_id,
                    province_id=row.province_id,
                    ward_id=row.ward_id or "",
                )
            )
            summary.skipped += 1
            summary.by_reason["province_not_found"] += 1
            continue

        if provinces[row.province_id] != row.country_id:
            issues.append(
                RowResult(
                    row_no=row.row_no,
                    internal_code=row.internal_code,
                    company_name=row.company_name,
                    status="skip",
                    reason="province_country_mismatch",
                    country_id=row.country_id,
                    province_id=row.province_id,
                    ward_id=row.ward_id or "",
                )
            )
            summary.skipped += 1
            summary.by_reason["province_country_mismatch"] += 1
            continue

        if row.ward_id:
            ward_info = wards.get(row.ward_id)
            if ward_info is None:
                issues.append(
                    RowResult(
                        row_no=row.row_no,
                        internal_code=row.internal_code,
                        company_name=row.company_name,
                        status="skip",
                        reason="ward_not_found",
                        country_id=row.country_id,
                        province_id=row.province_id,
                        ward_id=row.ward_id,
                    )
                )
                summary.skipped += 1
                summary.by_reason["ward_not_found"] += 1
                continue
            ward_province_id, ward_country_id = ward_info
            if ward_province_id != row.province_id or ward_country_id != row.country_id:
                issues.append(
                    RowResult(
                        row_no=row.row_no,
                        internal_code=row.internal_code,
                        company_name=row.company_name,
                        status="skip",
                        reason="ward_hierarchy_mismatch",
                        country_id=row.country_id,
                        province_id=row.province_id,
                        ward_id=row.ward_id,
                    )
                )
                summary.skipped += 1
                summary.by_reason["ward_hierarchy_mismatch"] += 1
                continue

        (
            client_id,
            cur_country,
            cur_province,
            cur_ward,
            cur_country_id,
            cur_province_id,
            cur_ward_id,
        ) = clients[row.internal_code]
        new_country_text = row.country
        new_province_text = row.province
        new_ward_text = row.ward
        new_country_id = row.country_id
        new_province_id = row.province_id
        new_ward_id = row.ward_id

        if (
            normalize_text_value(cur_country) == normalize_text_value(new_country_text)
            and normalize_text_value(cur_province) == normalize_text_value(new_province_text)
            and normalize_text_value(cur_ward) == normalize_text_value(new_ward_text)
            and normalize_uuid_value(cur_country_id) == normalize_uuid_value(new_country_id)
            and normalize_uuid_value(cur_province_id) == normalize_uuid_value(new_province_id)
            and normalize_uuid_value(cur_ward_id) == normalize_uuid_value(new_ward_id)
        ):
            summary.unchanged += 1
            continue

        updates.append(
            (
                client_id,
                new_country_text,
                new_province_text,
                new_ward_text,
                new_country_id,
                new_province_id,
                new_ward_id,
                row,
            )
        )
        summary.would_update += 1

    return updates, issues, summary


def apply_updates(
    cursor,
    updates: List[Tuple[str, str, str, Optional[str], str, str, Optional[str], ExcelRow]],
    commit_every: int,
) -> int:
    applied = 0
    sql = (
        "UPDATE dbo.client "
        "SET country = ?, province = ?, ward = ?, "
        "country_id = ?, province_id = ?, ward_id = ?, "
        "updated_at = SYSUTCDATETIME() "
        "WHERE client_id = ?"
    )

    for idx, (
        client_id,
        country,
        province,
        ward,
        country_id,
        province_id,
        ward_id,
        _row,
    ) in enumerate(updates, start=1):
        cursor.execute(
            sql,
            (country, province, ward, country_id, province_id, ward_id, client_id),
        )
        applied += 1
        if commit_every > 0 and idx % commit_every == 0:
            cursor.connection.commit()

    cursor.connection.commit()
    return applied


def default_report_path(xlsx_path: str) -> str:
    base_dir = os.path.dirname(os.path.abspath(xlsx_path))
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(base_dir, f"client_location_update_report_{ts}.csv")


def write_report(path: str, issues: List[RowResult]) -> None:
    if not issues:
        return
    fieldnames = [
        "row_no",
        "internal_code",
        "company_name",
        "status",
        "reason",
        "country_id",
        "province_id",
        "ward_id",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in issues:
            writer.writerow(
                {
                    "row_no": item.row_no,
                    "internal_code": item.internal_code,
                    "company_name": item.company_name,
                    "status": item.status,
                    "reason": item.reason,
                    "country_id": item.country_id,
                    "province_id": item.province_id,
                    "ward_id": item.ward_id,
                }
            )


def print_summary(summary: RunSummary, apply: bool, report_path: Optional[str], issue_count: int) -> None:
    mode = "APPLY" if apply else "DRY-RUN"
    print(f"\n=== {mode} SUMMARY ===")
    print(f"Excel rows processed: {summary.total_rows}")
    print(f"Would update:         {summary.would_update}")
    print(f"Unchanged:            {summary.unchanged}")
    print(f"Skipped:              {summary.skipped}")
    if apply:
        print(f"Applied:              {summary.applied}")
    if summary.by_reason:
        print("Skip reasons:")
        for reason, count in sorted(summary.by_reason.items()):
            print(f"  - {reason}: {count}")
    if issue_count and report_path:
        print(f"Report: {report_path}")
    if not apply and summary.would_update:
        print("\nChạy lại với --apply để ghi vào database.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Cập nhật 6 cột location (text + ID) cho client từ workbook Excel."
    )
    parser.add_argument(
        "--xlsx",
        default=DEFAULT_XLSX,
        help=f"Đường dẫn workbook (mặc định: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Ghi thay đổi vào database (mặc định: dry-run)",
    )
    parser.add_argument(
        "--commit-every",
        type=int,
        default=200,
        help="Commit mỗi N dòng khi --apply (0 = commit một lần cuối)",
    )
    parser.add_argument(
        "--report",
        default="",
        help="Đường dẫn CSV report cho dòng skip/lỗi (mặc định: cạnh file xlsx)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    xlsx_path = os.path.abspath(args.xlsx)
    if not os.path.isfile(xlsx_path):
        print(f"Không tìm thấy file Excel: {xlsx_path}")
        return 1

    conn = connect_sql_server(connection_string())
    try:
        cursor = conn.cursor()
        ensure_client_location_columns(cursor)

        countries, provinces, wards = load_lookup_maps(cursor)
        country_labels, province_labels, ward_labels = load_label_maps(cursor)
        clients = load_clients_by_internal_code(cursor)

        print(f"Đọc workbook: {xlsx_path}")
        excel_rows, excel_issues = read_excel_rows(
            xlsx_path, country_labels, province_labels, ward_labels
        )
        print(
            f"Đọc được {len(excel_rows)} dòng hợp lệ, "
            f"{len(excel_issues)} dòng lỗi/đọc sớm từ Excel"
        )

        updates, db_issues, summary = validate_and_plan_updates(
            excel_rows, clients, countries, provinces, wards
        )
        all_issues = excel_issues + db_issues

        if args.apply:
            if not updates:
                print("Không có dòng nào cần update.")
            else:
                summary.applied = apply_updates(cursor, updates, args.commit_every)
                print(f"Đã apply {summary.applied} dòng.")

        report_path = args.report.strip() or default_report_path(xlsx_path)
        if all_issues:
            write_report(report_path, all_issues)
        else:
            report_path = None

        print_summary(summary, args.apply, report_path, len(all_issues))
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
