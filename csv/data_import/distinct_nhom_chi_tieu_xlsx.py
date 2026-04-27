#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Liệt kê giá trị distinct cột "Nhóm Chỉ tiêu" trên Capability.xlsx (sheet Vietlabs + NTP).

Chạy mặc định (file ở repo Vietlabs/data/Capability.xlsx):
  python3 distinct_nhom_chi_tieu_xlsx.py

Tuỳ chọn:
  python3 distinct_nhom_chi_tieu_xlsx.py /đường/dẫn/Capability.xlsx
  python3 distinct_nhom_chi_tieu_xlsx.py --exclude-na
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Cần: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_XLSX = Path(__file__).resolve().parents[3] / "data" / "Capability.xlsx"


def norm_header(s) -> str:
    if s is None:
        return ""
    return " ".join(str(s).replace("\n", " ").split()).strip().lower()


def find_col_idx(header_row: tuple, want: str) -> int | None:
    target = norm_header(want)
    for j, cell in enumerate(header_row):
        if norm_header(cell) == target:
            return j
    for j, cell in enumerate(header_row):
        h = norm_header(cell)
        if "nhóm" in h and "chỉ tiêu" in h:
            return j
    return None


def main() -> None:
    p = argparse.ArgumentParser(description="Distinct cột Nhóm Chỉ tiêu từ Capability.xlsx")
    p.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Đường dẫn xlsx (mặc định: {DEFAULT_XLSX})",
    )
    p.add_argument("--exclude-na", action="store_true", help="Bỏ giá trị ô literal NA")
    p.add_argument("--out", type=Path, help="Ghi danh sách ra file UTF-8 (một dòng một giá trị)")
    args = p.parse_args()

    path: Path = args.xlsx
    if not path.is_file():
        print(f"Không tìm thấy file: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_vals: list[str] = []
    per_sheet: dict[str, list[str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            continue
        col_idx = find_col_idx(header_row, "Nhóm chỉ tiêu")
        if col_idx is None:
            print(f"[skip] Không có cột nhóm chỉ tiêu: sheet={sheet_name!r}", file=sys.stderr)
            continue
        vals: list[str] = []
        for row in rows:
            if row is None or col_idx >= len(row):
                continue
            v = row[col_idx]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            vals.append(s)
        per_sheet[sheet_name] = vals
        all_vals.extend(vals)

    wb.close()

    distinct = sorted(set(all_vals), key=lambda x: (x.upper(), x))
    if args.exclude_na:
        distinct = [x for x in distinct if x.strip().upper() != "NA"]

    print(f"File: {path}")
    print(f"Số giá trị distinct: {len(distinct)}")
    for i, v in enumerate(distinct, 1):
        print(f"{i:3}. {v}")

    print("\nTheo sheet:")
    for sn, vals in per_sheet.items():
        d = sorted(set(vals), key=lambda x: (x.upper(), x))
        if args.exclude_na:
            d = [x for x in d if x.strip().upper() != "NA"]
        print(f"  {sn}: {len(d)} distinct / {len(vals)} ô có giá trị")

    if args.out:
        args.out.write_text("\n".join(distinct) + "\n", encoding="utf-8")
        print(f"\nĐã ghi: {args.out}")


if __name__ == "__main__":
    main()
