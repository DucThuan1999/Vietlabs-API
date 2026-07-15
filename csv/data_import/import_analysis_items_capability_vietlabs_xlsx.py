#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import analysis_item từ Capability.xlsx, sheet Vietlabs và/hoặc NTP.

(1) analysis_item_code = giá trị cột "Mã chỉ tiêu" (định dạng CT-...), không tự sinh CT-0001.
(2) FK tra master: equipment_type, analysis_group, sample_matrix / sample_matrix_group;
    standard, reference_method, unit_of_measure, laboratory_technique khi có dữ liệu cột tương ứng.

Sheet NTP: giá lấy từ cột "Đơn giá (NTP)"; kỹ thuật từ "Kỹ thuật"; ghi chú từ "Ghi chú (Tên NTP)".
Cột INSERT/UPDATE được lọc theo INFORMATION_SCHEMA (tương thích DB cũ/mới).
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Set, Tuple

# Cùng thư mục với import_analysis_item
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import import_analysis_item as iai
import excel_cell_tiptap as ect

try:
    import pyodbc
except ImportError:
    pyodbc = None  # type: ignore

try:
    import openpyxl
except ImportError:
    print("Can cai: pip install openpyxl")
    sys.exit(1)

DEFAULT_XLSX = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "Danh mục Năng lực v2.xlsx")
)
# Legacy fallback
DEFAULT_XLSX_LEGACY = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "Capability.xlsx")
)

def norm_header_cell(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).replace("\n", " ").replace('"', "").replace("'", "").strip()
    s = re.sub(r"\s+", " ", s).lower()
    return s


def build_column_map(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    """Map logical field -> column index from row 1."""
    pairs: List[Tuple[str, int]] = []
    for i, v in enumerate(header_row):
        pairs.append((norm_header_cell(v), i))

    def find(pred) -> Optional[int]:
        for nh, idx in pairs:
            if nh and pred(nh):
                return idx
        return None

    m: Dict[str, int] = {}
    code_i = find(lambda nh: "mã chỉ tiêu" in nh)
    if code_i is not None:
        m["code"] = code_i
    smg = find(lambda nh: nh == "nhóm nền mẫu")
    if smg is not None:
        m["sm_group"] = smg
    sm = find(lambda nh: nh == "nền mẫu")
    if sm is not None:
        m["sm"] = sm
    std = find(
        lambda nh: "tiêu chuẩn" in nh
        and ("qui chuẩn" in nh or "quy chuẩn" in nh)
    )
    if std is not None:
        m["standard"] = std
    ag = find(lambda nh: nh == "nhóm chỉ tiêu")
    if ag is not None:
        m["analysis_group"] = ag
    nv = find(lambda nh: nh == "tên chỉ tiêu")
    if nv is not None:
        m["name_vi"] = nv
    ne = find(lambda nh: nh == "tên tiếng anh")
    if ne is not None:
        m["name_en"] = ne
    sn = find(lambda nh: nh == "tên viết tắt")
    if sn is not None:
        m["short_name"] = sn
    rm = find(lambda nh: nh == "phương pháp")
    if rm is not None:
        m["reference_method"] = rm
    sv = find(lambda nh: nh == "giá trị")
    if sv is not None:
        m["standard_value"] = sv
    lod = find(lambda nh: "lod" in nh and "giới hạn phát hiện" in nh)
    if lod is None:
        lod = find(lambda nh: "lod" in nh and "loq" not in nh)
    if lod is not None:
        m["lod"] = lod
    loq = find(lambda nh: "loq" in nh)
    if loq is not None:
        m["loq"] = loq
    uom = find(lambda nh: nh == "đvt" or "đơn vị tính" in nh)
    if uom is not None:
        m["uom"] = uom
    eq = find(lambda nh: "thiết bị" in nh or "equipment" in nh)
    if eq is not None:
        m["equipment"] = eq
    t1 = find(lambda nh: "tat thường" in nh or ("tat" in nh and "thường" in nh))
    if t1 is not None:
        m["tat_normal"] = t1
    t2 = find(lambda nh: "tat nhanh" in nh)
    if t2 is not None:
        m["tat_fast"] = t2
    t3 = find(lambda nh: "tat khẩn" in nh)
    if t3 is not None:
        m["tat_urgent"] = t3
    lt = find(
        lambda nh: "bộ phận phụ trách" in nh or ("kỹ thuật" in nh and "bộ phận" in nh)
    )
    if lt is None:
        lt = find(lambda nh: nh == "kỹ thuật")
    if lt is not None:
        m["lab_technique"] = lt
    sq = find(lambda nh: nh == "khối lượng tiêu chuẩn")
    if sq is not None:
        m["std_qty"] = sq
    squ = find(lambda nh: "đvt khối lượng" in nh)
    if squ is not None:
        m["std_qty_uom"] = squ
    pn = find(lambda nh: "đơn giá chuẩn_new" in nh or nh == "đơn giá chuẩn_new")
    if pn is None:
        pn = find(lambda nh: "đơn giá" in nh and "_new" in nh)
    if pn is None:
        pn = find(
            lambda nh: nh == "đơn giá chuẩn"
            or ("đơn giá chuẩn" in nh and "_new" not in nh and "_old" not in nh)
        )
    if pn is None:
        pn = find(lambda nh: "đơn giá" in nh and "ntp" in nh)
    if pn is not None:
        m["unit_price"] = pn
    gn = find(lambda nh: "giá nhóm chuẩn_new" in nh or nh == "giá nhóm chuẩn_new")
    if gn is None:
        gn = find(lambda nh: nh == "giá nhóm chuẩn")
    if gn is None:
        gn = find(lambda nh: "giá nhóm" in nh and "_new" in nh)
    if gn is not None:
        m["group_price"] = gn
    st = find(lambda nh: nh == "trạng thái")
    if st is not None:
        m["status"] = st
    notes = find(lambda nh: "ghi chú" in nh)
    if notes is not None:
        m["notes"] = notes
    return m


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def is_valid_ct_code(code: str) -> bool:
    if not code:
        return False
    s = code.strip()
    return bool(re.match(r"(?i)^CT-\S", s))


def is_ntp_sheet(sheet_name: str) -> bool:
    return str(sheet_name or "").strip().upper().startswith("NTP")


def is_blank_sentinel(s: str) -> bool:
    return iai.is_blank_analysis_group_cell(s)


def parse_lod_loq_safe(raw: str) -> Tuple[Optional[float], Optional[str]]:
    """Parse LOD/LOQ; tra ve (value, warning) — khong rut gon text pham vi do."""
    if not raw or is_blank_sentinel(raw):
        return None, None
    low = raw.strip().lower()
    if any(tok in low for tok in ("phạm vi", "pham vi", ">=", "<=", ">", "<")):
        return None, f"LOD/LOQ khong parse an toan: {raw!r}"
    val, _u = iai.parse_decimal_with_unit(raw)
    return val, None


def _tat_cell_sentinel_none(s: str) -> bool:
    u = s.strip().upper().replace(" ", "")
    return u in ("", "NA", "N/A", "-", "--", "NONE", "NULL", "KHÔNG", "KHONGCO")


def _tat_first_number(s: str) -> Optional[float]:
    m = re.search(r"(\d+(?:[.,]\d+)?)", s.replace(" ", ""))
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


def parse_tat_cell_to_hours(
    value: Any,
    header_norm: str,
    sheet_name: str,
) -> Optional[int]:
    """
    Chuẩn hóa TAT về số giờ cho DB/frontend (tat_unit = Hours).

    - Excel NTP: thường là \"10 ngày\", \"7 ngày\" hoặc NA.
    - Excel Vietlabs: header có (giờ), ô thường là số giờ.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value != int(value):
            x = float(value)
        else:
            x = float(int(value))
        if x <= 0:
            return None
        if "giờ" in header_norm or "gio" in header_norm or "hour" in header_norm:
            return int(x)
        if "ngày" in header_norm or "ngay" in header_norm or "day" in header_norm:
            return int(x * 24)
        if is_ntp_sheet(sheet_name):
            return int(x * 24)
        return int(x)

    s = str(value).strip()
    if not s or _tat_cell_sentinel_none(s):
        return None

    low = s.lower()
    if "ngày" in low or "ngay" in low:
        n = _tat_first_number(s)
        if n is None or n <= 0:
            return None
        return int(round(n * 24))
    if "giờ" in low or "gio" in low or re.search(r"\bh\b", low) or "hour" in low:
        n = _tat_first_number(s)
        if n is None or n <= 0:
            return None
        return int(n) if n == int(n) else int(n)

    n = _tat_first_number(s)
    if n is None:
        return None
    if n <= 0:
        return None
    if "giờ" in header_norm or "gio" in header_norm or "hour" in header_norm:
        return int(n) if n == int(n) else int(n)
    if "ngày" in header_norm or "ngay" in header_norm or "day" in header_norm:
        return int(round(n * 24))
    if is_ntp_sheet(sheet_name):
        return int(round(n * 24))
    return int(n) if n == int(n) else int(n)


def is_transient_odbc_error(exc: BaseException) -> bool:
    """Lỗi kết nối / timeout thường gặp khi import lâu (08S01, HYT00, ...)."""
    if pyodbc is None or not isinstance(exc, pyodbc.Error):
        return False
    args = getattr(exc, "args", ()) or ()
    if args:
        code = str(args[0]).upper()
        if code in ("08S01", "HYT00", "08001", "HY000"):
            return True
    msg = str(exc).lower()
    if "communication link failure" in msg:
        return True
    if "connection may have been terminated" in msg:
        return True
    if "timeout" in msg and ("expired" in msg or "hết hạn" in msg):
        return True
    return False


def reconnect_import_ctx(ctx: Dict[str, Any]) -> None:
    """Đóng kết nối cũ, mở lại và nạp lại mappings (sau 08S01 / link failure)."""
    old = ctx.get("connection")
    if old is not None:
        try:
            old.close()
        except Exception:
            pass
    conn = iai.pyodbc.connect(iai.CONNECTION_STRING)
    ctx["connection"] = conn
    ctx["table_cols"] = get_table_columns(conn, "analysis_item")
    ctx["mappings"] = iai.load_mappings(conn)
    augment_master_maps(conn, ctx["mappings"])


def get_table_columns(connection, table: str, table_schema: str = "dbo") -> Set[str]:
    """Chỉ lấy cột của schema chỉ định — tránh gộp nhầm bảng trùng tên khác schema (vd. cột `unit` không tồn tại trên dbo)."""
    cur = connection.cursor()
    cur.execute(
        """
        SELECT LOWER(COLUMN_NAME) FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """,
        table_schema,
        table,
    )
    return {r[0] for r in cur.fetchall()}


def _master_lookup_key(label: str) -> str:
    key = iai.normalize_text(label)
    return key if key else label.strip().casefold()


def augment_master_maps(connection, mappings: Dict) -> None:
    """Thêm map standard / reference_method / unit_of_measure / laboratory_technique."""
    cur = connection.cursor()
    mappings.setdefault("standards", {})
    mappings.setdefault("reference_methods", {})
    mappings.setdefault("unit_of_measures", {})
    mappings.setdefault("laboratory_techniques", {})

    loaders = [
        (
            "standards",
            "SELECT standard_id, name_vi, name_en, standard_code FROM standard",
            (1, 2, 3),
        ),
        (
            "reference_methods",
            "SELECT reference_method_id, name_vi, name_en, reference_method_code FROM reference_method",
            (1, 2, 3),
        ),
        (
            "unit_of_measures",
            "SELECT unit_of_measure_id, name_vi, name_en, unit_of_measure_code FROM unit_of_measure",
            (1, 2, 3),
        ),
        (
            "laboratory_techniques",
            "SELECT laboratory_technique_id, name_vi, name_en, technique_code FROM laboratory_technique",
            (1, 2, 3),
        ),
    ]
    for key, sql, text_cols in loaders:
        try:
            cur.execute(sql)
            for row in cur.fetchall():
                rid = row[0]
                for j in text_cols:
                    label = row[j]
                    if label and str(label).strip():
                        lbl = str(label).strip()
                        mappings[key][iai.normalize_text(lbl)] = rid
                        alt = _master_lookup_key(lbl)
                        if alt:
                            mappings[key][alt] = rid
        except Exception:
            # Bảng không tồn tại trên DB cũ
            mappings[key] = {}


def resolve_master(mappings: Dict, bucket: str, text: str) -> Optional[str]:
    if not text or not str(text).strip() or is_blank_sentinel(str(text)):
        return None
    b = mappings.get(bucket) or {}
    return b.get(_master_lookup_key(str(text).strip()))


def _next_prefixed_code(cur, table: str, code_col: str, prefix: str) -> str:
    cur.execute(
        f"""SELECT MAX(CAST(SUBSTRING({code_col}, LEN(?) + 1, 20) AS INT))
        FROM {table} WHERE {code_col} LIKE ? + '%'""",
        prefix,
        prefix,
    )
    row = cur.fetchone()
    n = int(row[0]) if row and row[0] is not None else 0
    return f"{prefix}{n + 1:03d}"


def get_or_create_standard(connection, name: str, mappings: Dict) -> Optional[str]:
    if not name or not str(name).strip() or is_blank_sentinel(str(name)):
        return None
    label = str(name).strip()
    key = _master_lookup_key(label)
    bucket = mappings.setdefault("standards", {})
    if key in bucket:
        return str(bucket[key])
    cur = connection.cursor()
    cur.execute(
        "SELECT standard_id FROM standard WHERE LTRIM(RTRIM(name_vi)) = ?",
        label,
    )
    ex = cur.fetchone()
    if ex:
        sid = str(ex[0])
        bucket[key] = sid
        return sid
    sid = str(uuid.uuid4())
    code = _next_prefixed_code(cur, "standard", "standard_code", "TC-")
    now = datetime.now(timezone.utc)
    cur.execute(
        """INSERT INTO standard (standard_id, standard_code, name_vi, name_en, status, created_at)
        VALUES (?, ?, ?, ?, N'Active', ?)""",
        (sid, code, label, label, now),
    )
    connection.commit()
    bucket[key] = sid
    return sid


def get_or_create_reference_method(connection, name: str, mappings: Dict) -> Optional[str]:
    if not name or not str(name).strip() or is_blank_sentinel(str(name)):
        return None
    label = str(name).strip()
    key = _master_lookup_key(label)
    bucket = mappings.setdefault("reference_methods", {})
    if key in bucket:
        return str(bucket[key])
    cur = connection.cursor()
    cur.execute(
        "SELECT reference_method_id FROM reference_method WHERE LTRIM(RTRIM(name_vi)) = ?",
        label,
    )
    ex = cur.fetchone()
    if ex:
        rid = str(ex[0])
        bucket[key] = rid
        return rid
    rid = str(uuid.uuid4())
    code = _next_prefixed_code(cur, "reference_method", "reference_method_code", "PP-")
    now = datetime.now(timezone.utc)
    cur.execute(
        """INSERT INTO reference_method (
            reference_method_id, reference_method_code, name_vi, name_en, status, created_at
        ) VALUES (?, ?, ?, ?, N'Active', ?)""",
        (rid, code, label, label, now),
    )
    connection.commit()
    bucket[key] = rid
    return rid


def get_or_create_unit_of_measure(connection, name: str, mappings: Dict) -> Optional[str]:
    if not name or not str(name).strip() or is_blank_sentinel(str(name)):
        return None
    label = str(name).strip()
    key = _master_lookup_key(label)
    bucket = mappings.setdefault("unit_of_measures", {})
    if key in bucket:
        return str(bucket[key])
    cur = connection.cursor()
    cur.execute(
        "SELECT unit_of_measure_id FROM unit_of_measure WHERE LTRIM(RTRIM(name_vi)) = ?",
        label,
    )
    ex = cur.fetchone()
    if ex:
        uid = str(ex[0])
        bucket[key] = uid
        return uid
    uid = str(uuid.uuid4())
    code = _next_prefixed_code(cur, "unit_of_measure", "unit_of_measure_code", "DVT-")
    now = datetime.now(timezone.utc)
    cur.execute(
        """INSERT INTO unit_of_measure (
            unit_of_measure_id, unit_of_measure_code, name_vi, name_en, status, created_at
        ) VALUES (?, ?, ?, ?, N'Active', ?)""",
        (uid, code, label, label, now),
    )
    connection.commit()
    bucket[key] = uid
    return uid


# (py_key, sql_column) — chỉ ghi nếu cột có trong DB
ROW_SQL_FIELDS = [
    ("name_vi", "name_vi"),
    ("name_en", "name_en"),
    ("short_name", "short_name"),
    ("display_name_vi", "display_name_vi"),
    ("display_name_en", "display_name_en"),
    ("display_short_name", "display_short_name"),
    ("equipment_type_id", "equipment_type_id"),
    ("analysis_group_id", "analysis_group_id"),
    ("sample_matrix_id", "sample_matrix_id"),
    ("sample_matrix_group_id", "sample_matrix_group_id"),
    ("reference_method_id", "reference_method_id"),
    ("standard_id", "standard_id"),
    ("unit_of_measure_id", "unit_of_measure_id"),
    ("laboratory_technique_id", "laboratory_technique_id"),
    ("published_group_code", "published_group_code"),
    ("lod", "lod"),
    ("loq", "loq"),
    ("standard_value", "standard_value"),
    ("standard_quantity_text", "standard_quantity_text"),
    ("standard_quantity_unit_of_measure_id", "standard_quantity_unit_of_measure_id"),
    ("unit_price", "unit_price"),
    ("status", "status"),
    ("notes", "notes"),
]


def filter_row_for_table(row: Dict[str, Any], table_cols: Set[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for pyk, sqlc in ROW_SQL_FIELDS:
        if sqlc not in table_cols:
            continue
        if pyk not in row:
            continue
        val = row[pyk]
        if val is None and pyk in (
            "equipment_type_id",
            "reference_method_id",
            "standard_id",
            "unit_of_measure_id",
            "laboratory_technique_id",
            "published_group_code",
            "short_name",
            "display_name_vi",
            "display_name_en",
            "display_short_name",
            "standard_value",
            "standard_quantity_text",
            "standard_quantity_unit_of_measure_id",
            "lod",
            "loq",
            "unit_price",
            "notes",
            "name_en",
        ):
            out[sqlc] = None
            continue
        if val is None:
            continue
        out[sqlc] = val
    return out


def upsert_analysis_item(
    connection,
    table_cols: Set[str],
    analysis_item_code: str,
    row_values: Dict[str, Any],
    dry_run: bool,
) -> Tuple[Optional[str], str]:
    """
    Returns (analysis_item_id, action) action in insert|update|skip|error
    """
    cursor = connection.cursor()
    cursor.execute(
        "SELECT analysis_item_id FROM analysis_item WHERE analysis_item_code = ?",
        analysis_item_code,
    )
    ex = cursor.fetchone()
    now = datetime.now(timezone.utc)
    filtered = filter_row_for_table(row_values, table_cols)

    if ex:
        aid = str(ex[0])
        if dry_run:
            return aid, "update"
        sets = [f"{k} = ?" for k in filtered.keys()]
        sets.append("updated_at = ?")
        vals = list(filtered.values()) + [now, aid]
        if sets:
            sql = f"UPDATE analysis_item SET {', '.join(sets)} WHERE analysis_item_id = ?"
            cursor.execute(sql, vals)
        return aid, "update"

    aid = str(uuid.uuid4())
    if dry_run:
        return aid, "insert"

    insert_cols = ["analysis_item_id", "analysis_item_code"] + list(filtered.keys())
    if "created_at" in table_cols:
        insert_cols.append("created_at")
    if "updated_at" in table_cols:
        insert_cols.append("updated_at")

    placeholders = ", ".join(["?"] * len(insert_cols))
    values: List[Any] = [aid, analysis_item_code]
    for k in filtered.keys():
        values.append(filtered[k])
    if "created_at" in table_cols:
        values.append(now)
    if "updated_at" in table_cols:
        values.append(None)

    col_sql = ", ".join(insert_cols)
    sql = f"INSERT INTO analysis_item ({col_sql}) VALUES ({placeholders})"
    cursor.execute(sql, values)
    return aid, "insert"


def apply_tat_and_group_price(
    connection,
    analysis_item_id: str,
    analysis_group_id: Optional[str],
    tat_normal: Optional[int],
    tat_fast: Optional[int],
    tat_urgent: Optional[int],
    whole_group_raw: str,
    dry_run: bool,
) -> None:
    if dry_run:
        return
    cursor = connection.cursor()
    for tat_type, tat_value in [
        ("Normal", tat_normal),
        ("Fast", tat_fast),
        ("Urgent", tat_urgent),
    ]:
        if tat_value is not None and tat_value > 0:
            cursor.execute(
                """
                SELECT analysis_item_tat_id FROM analysis_item_tat
                WHERE analysis_item_id = ? AND tat_type = ?
                """,
                analysis_item_id,
                tat_type,
            )
            existing = cursor.fetchone()
            now = datetime.now(timezone.utc)
            if existing:
                cursor.execute(
                    """
                    UPDATE analysis_item_tat SET tat_value = ?, updated_at = ?
                    WHERE analysis_item_id = ? AND tat_type = ?
                    """,
                    tat_value,
                    now,
                    analysis_item_id,
                    tat_type,
                )
            else:
                tid = str(uuid.uuid4())
                cursor.execute(
                    """
                    INSERT INTO analysis_item_tat (
                        analysis_item_tat_id, analysis_item_id, tat_type,
                        tat_value, tat_unit, created_at
                    ) VALUES (?, ?, ?, ?, 'Hours', ?)
                    """,
                    tid,
                    analysis_item_id,
                    tat_type,
                    tat_value,
                    now,
                )

    if whole_group_raw and analysis_group_id:
        price = iai.parse_price(whole_group_raw)
        if price:
            try:
                cursor.execute(
                    """
                    UPDATE analysis_group SET whole_group_standard_price = ?
                    WHERE analysis_group_id = ?
                    """,
                    price,
                    analysis_group_id,
                )
            except Exception:
                pass


def process_workbook(
    xlsx_path: str,
    ctx: Dict[str, Any],
    dry_run: bool,
    sheet_name: str,
    max_rows: Optional[int] = None,
    only_codes: Optional[List[str]] = None,
) -> Dict[str, int]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True, rich_text=True)
    if sheet_name not in wb.sheetnames:
        print(f"Loi: Khong co sheet {sheet_name!r}")
        wb.close()
        return {"insert": 0, "update": 0, "skip": 0, "error": 0}
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        print(f"Sheet {sheet_name} rong")
        wb.close()
        return {"insert": 0, "update": 0, "skip": 0, "error": 0}

    colmap = build_column_map(tuple(header_row))
    required = ("code", "name_vi", "analysis_group", "equipment", "sm_group", "sm")
    missing = [k for k in required if k not in colmap]
    if missing:
        print(f"Loi sheet {sheet_name}: Thieu cot (header): {missing}")
        wb.close()
        return {"insert": 0, "update": 0, "skip": 0, "error": 0}

    stats = {"insert": 0, "update": 0, "skip": 0, "error": 0}
    errors: List[str] = []
    only_cf: Optional[Set[str]] = None
    if only_codes:
        only_cf = {c.strip().casefold() for c in only_codes if c and str(c).strip()}
        print(f"  Chi xu ly {len(only_cf)} ma chi tieu")
    if max_rows is not None:
        print(f"  Gioi han: toi da {max_rows} dong insert+update thanh cong")

    max_row = ws.max_row or 1
    for ridx in range(2, max_row + 1):
        if max_rows is not None and (stats["insert"] + stats["update"]) >= max_rows:
            break

        def g(key: str) -> str:
            i = colmap.get(key)
            if i is None:
                return ""
            cell = ws.cell(ridx, i + 1)
            return cell_str(cell.value)

        code = g("code")
        if not code:
            stats["skip"] += 1
            continue
        if only_cf is not None and code.strip().casefold() not in only_cf:
            continue
        if not is_valid_ct_code(code):
            errors.append(f"Dong {ridx}: ma khong hop le (can CT-...): {code!r}")
            stats["error"] += 1
            continue

        name_vi_cell = ws.cell(ridx, colmap["name_vi"] + 1)
        name_vi_raw = cell_str(name_vi_cell.value)
        if not name_vi_raw:
            stats["skip"] += 1
            continue

        name_vi_plain, display_name_vi = ect.cell_to_plain_and_display(name_vi_cell)
        name_en_plain, display_name_en = None, None
        if "name_en" in colmap:
            name_en_cell = ws.cell(ridx, colmap["name_en"] + 1)
            name_en_plain, display_name_en = ect.cell_to_plain_and_display(name_en_cell)
        short_plain, display_short_name = None, None
        if "short_name" in colmap:
            short_cell = ws.cell(ridx, colmap["short_name"] + 1)
            short_plain, display_short_name = ect.cell_to_plain_and_display(short_cell)

        sm_group_val = g("sm_group")
        sm_val = g("sm")
        if not sm_group_val or is_blank_sentinel(sm_group_val):
            errors.append(f"Dong {ridx} {code}: thieu Nhom nen mau")
            stats["error"] += 1
            continue
        if not sm_val or is_blank_sentinel(sm_val):
            errors.append(f"Dong {ridx} {code}: thieu Nen mau")
            stats["error"] += 1
            continue

        name_vi = name_vi_plain or name_vi_raw.strip()
        name_en_raw = g("name_en")
        short_raw = g("short_name")
        name_en = name_en_plain or (
            name_en_raw if name_en_raw and not is_blank_sentinel(name_en_raw) else name_vi
        )
        short_name = short_plain or (
            short_raw if short_raw and not is_blank_sentinel(short_raw) else None
        )
        ag_name = g("analysis_group")
        eq_raw = g("equipment")
        eq_name = None if is_blank_sentinel(eq_raw) else eq_raw
        ag_blank = iai.is_blank_analysis_group_cell(ag_name)

        std_text = g("standard")
        if is_blank_sentinel(std_text):
            std_text = ""
        ref_text = g("reference_method")
        if is_blank_sentinel(ref_text):
            ref_text = ""
        uom_text = g("uom")
        if is_blank_sentinel(uom_text):
            uom_text = ""
        lab_text = g("lab_technique")
        if is_blank_sentinel(lab_text):
            lab_text = ""

        lod_raw = g("lod")
        loq_raw = g("loq")
        lod_value, lod_warn = parse_lod_loq_safe(lod_raw)
        if lod_warn:
            errors.append(f"Dong {ridx} {code}: {lod_warn}")
        loq_value, loq_warn = parse_lod_loq_safe(loq_raw)
        if loq_warn:
            errors.append(f"Dong {ridx} {code}: {loq_warn}")

        std_val_cell = g("standard_value")
        standard_value = (
            std_val_cell if std_val_cell and not is_blank_sentinel(std_val_cell) else None
        )

        sq_text = g("std_qty")
        standard_quantity_text = sq_text if sq_text else None
        squ_text = g("std_qty_uom")

        up_raw = g("unit_price")
        if up_raw and not is_blank_sentinel(up_raw):
            unit_price = iai.parse_price(up_raw)
            if unit_price is None:
                errors.append(f"Dong {ridx} {code}: Don gia khong hop le: {up_raw!r}")
        else:
            unit_price = None

        st = g("status")
        if st and "inactive" in st.lower():
            status = "Inactive"
        else:
            status = "Active"

        notes_raw = g("notes")
        notes = notes_raw if notes_raw and not is_blank_sentinel(notes_raw) else None

        hdr = tuple(header_row)

        def tat_from_col(key: str) -> Optional[int]:
            idx = colmap.get(key)
            if idx is None:
                return None
            hn = norm_header_cell(hdr[idx]) if idx < len(hdr) else ""
            cell = ws.cell(ridx, idx + 1)
            return parse_tat_cell_to_hours(cell.value, hn, sheet_name)

        tat_n = tat_from_col("tat_normal")
        tat_f = tat_from_col("tat_fast")
        tat_u = tat_from_col("tat_urgent")

        group_price = ""
        if "group_price" in colmap:
            group_price = cell_str(ws.cell(ridx, colmap["group_price"] + 1).value)

        max_attempts = 5
        for attempt in range(max_attempts):
            connection = ctx["connection"]
            mappings = ctx["mappings"]
            table_cols = ctx["table_cols"]
            try:
                if ag_blank:
                    analysis_group_id = None
                else:
                    analysis_group_id = iai.get_or_create_analysis_group(
                        connection, ag_name, mappings
                    )
                equipment_type_id = (
                    iai.get_or_create_equipment_type(connection, eq_name, mappings)
                    if eq_name
                    else None
                )
                sm_group_id = iai.get_or_create_sample_matrix_group(
                    connection, g("sm_group"), mappings
                )
                sm_id = iai.get_or_create_sample_matrix(
                    connection, g("sm"), sm_group_id, mappings
                )

                standard_id = get_or_create_standard(connection, std_text, mappings) if std_text else None
                reference_method_id = (
                    get_or_create_reference_method(connection, ref_text, mappings)
                    if ref_text
                    else None
                )
                unit_of_measure_id = (
                    get_or_create_unit_of_measure(connection, uom_text, mappings)
                    if uom_text
                    else None
                )
                laboratory_technique_id = (
                    resolve_master(mappings, "laboratory_techniques", lab_text)
                    if lab_text
                    else None
                )
                squ_id = (
                    get_or_create_unit_of_measure(connection, squ_text, mappings)
                    if squ_text
                    else None
                )

                row_payload: Dict[str, Any] = {
                    "name_vi": name_vi,
                    "name_en": name_en,
                    "short_name": short_name,
                    "display_name_vi": display_name_vi,
                    "display_name_en": display_name_en,
                    "display_short_name": display_short_name,
                    "equipment_type_id": equipment_type_id,
                    "analysis_group_id": analysis_group_id,
                    "sample_matrix_id": sm_id,
                    "sample_matrix_group_id": sm_group_id,
                    "reference_method_id": reference_method_id,
                    "standard_id": standard_id,
                    "unit_of_measure_id": unit_of_measure_id,
                    "laboratory_technique_id": laboratory_technique_id,
                    "published_group_code": None,
                    "lod": lod_value,
                    "loq": loq_value,
                    "standard_value": standard_value,
                    "standard_quantity_text": standard_quantity_text,
                    "standard_quantity_unit_of_measure_id": squ_id,
                    "unit_price": unit_price,
                    "status": status,
                    "notes": notes,
                }

                aid, action = upsert_analysis_item(
                    connection, table_cols, code.strip(), row_payload, dry_run
                )
                stats[action] = stats.get(action, 0) + 1
                if aid:
                    apply_tat_and_group_price(
                        connection,
                        aid,
                        analysis_group_id,
                        tat_n,
                        tat_f,
                        tat_u,
                        group_price,
                        dry_run,
                    )
                if not dry_run:
                    connection.commit()
                break
            except Exception as e:
                if pyodbc and isinstance(e, pyodbc.Error) and is_transient_odbc_error(e):
                    if attempt + 1 < max_attempts:
                        print(
                            f"  Canh bao: loi ket noi tam thoi dong {ridx} "
                            f"({e!s}), thu lai {attempt + 2}/{max_attempts}..."
                        )
                        try:
                            connection.rollback()
                        except Exception:
                            pass
                        reconnect_import_ctx(ctx)
                        continue
                stats["error"] += 1
                errors.append(f"Dong {ridx} {code}: {e}")
                try:
                    connection.rollback()
                except Exception:
                    pass
                break

    wb.close()

    print(f"\nKet qua sheet {sheet_name}:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    if errors:
        print("\nCanh bao / loi (toi da 20):")
        for e in errors[:20]:
            print(f"  - {e}")
        if len(errors) > 20:
            print(f"  ... va {len(errors) - 20} dong nua")
    return stats


def main():
    parser = argparse.ArgumentParser(
        description="Import analysis_item tu Capability.xlsx (Vietlabs va/hoac NTP)"
    )
    parser.add_argument(
        "--xlsx",
        default=DEFAULT_XLSX if os.path.isfile(DEFAULT_XLSX) else None,
        help="Duong dan Capability.xlsx",
    )
    parser.add_argument(
        "--sheet",
        choices=("vietlabs", "ntp", "ntp_bo_sung", "all"),
        default="vietlabs",
        help="Sheet: vietlabs (mac dinh), ntp, ntp_bo_sung, hoac all",
    )
    parser.add_argument("--dry-run", action="store_true", help="Khong ghi DB")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=None,
        metavar="N",
        help="Dung sau N lan insert+update thanh cong (de thu nghiem)",
    )
    parser.add_argument(
        "--only-code",
        action="append",
        dest="only_codes",
        metavar="CODE",
        help="Chi xu ly ma chi tieu (lap lai tuy chon de nhieu ma)",
    )
    args = parser.parse_args()

    if not args.xlsx or not os.path.isfile(args.xlsx):
        if os.path.isfile(DEFAULT_XLSX):
            args.xlsx = DEFAULT_XLSX
        elif os.path.isfile(DEFAULT_XLSX_LEGACY):
            args.xlsx = DEFAULT_XLSX_LEGACY
    if not args.xlsx or not os.path.isfile(args.xlsx):
        print("Loi: Chi dinh --xlsx hop le (mac dinh: data/Danh muc Nang luc v2.xlsx)")
        sys.exit(1)

    if sys.platform == "win32":
        import io

        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    print("=" * 60)
    print("Import analysis_item tu Capability.xlsx")
    print("=" * 60)

    import openpyxl as _ox

    wb_probe = _ox.load_workbook(args.xlsx, read_only=True)
    available = set(wb_probe.sheetnames)
    wb_probe.close()

    sheets: List[str]
    if args.sheet == "all":
        sheets = ["Vietlabs"]
        for sn in ("NTP", "NTP bổ sung"):
            if sn in available:
                sheets.append(sn)
    elif args.sheet == "ntp":
        sheets = ["NTP"] if "NTP" in available else []
    elif args.sheet == "ntp_bo_sung":
        sheets = ["NTP bổ sung"] if "NTP bổ sung" in available else []
    else:
        sheets = ["Vietlabs"]

    ctx: Dict[str, Any] = {"connection": iai.pyodbc.connect(iai.CONNECTION_STRING)}
    try:
        conn = ctx["connection"]
        ctx["table_cols"] = get_table_columns(conn, "analysis_item")
        ctx["mappings"] = iai.load_mappings(conn)
        augment_master_maps(conn, ctx["mappings"])
        print(f"  analysis_item columns: {len(ctx['table_cols'])}")
        print(f"  standards: {len(ctx['mappings'].get('standards') or {})}")
        print(f"  reference_methods: {len(ctx['mappings'].get('reference_methods') or {})}")
        print(f"  unit_of_measures: {len(ctx['mappings'].get('unit_of_measures') or {})}")
        print(f"  laboratory_techniques: {len(ctx['mappings'].get('laboratory_techniques') or {})}")
        print(f"  Sheets: {', '.join(sheets)}")
        for sn in sheets:
            print(f"\n--- {sn} ---")
            process_workbook(
                args.xlsx,
                ctx,
                args.dry_run,
                sn,
                max_rows=args.max_rows,
                only_codes=args.only_codes,
            )
    finally:
        c = ctx.get("connection")
        if c is not None:
            try:
                c.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()
