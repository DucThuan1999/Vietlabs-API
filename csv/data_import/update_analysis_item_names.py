#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cập nhật tên chỉ tiêu (analysis_item.name_vi / name_en) từ CSV hoặc từ Capability.xlsx.

- CSV: áp kiểu hoa–thường theo tên đang có trên DB (hàm preserve_case_from_reference).
- Capability.xlsx (--xlsx): ghi đúng chuỗi name_vi / name_en như trong ô Excel (không đổi hoa thường).

Chỉ ghi các cột: name_vi, name_en, updated_at. Không cập nhật analysis_item_code
hay bất kỳ cột nào khác. Mã chỉ dùng để tìm đúng bản ghi (SELECT).

Capability.xlsx (--xlsx): đọc cột "Mã chỉ tiêu", "Tên chỉ tiêu", "Tên tiếng anh"
trên các sheet (mặc định Vietlabs rồi NTP). Mã trùng giữ bản ghi gặp trước theo thứ tự sheet.

CSV (UTF-8 BOM hoặc không):
  - Bắt buộc: mã chỉ tiêu — một trong các tên cột:
      analysis_item_code | mã chỉ tiêu | ma_chi_tieu
  - Tên mới (ít nhất một cột):
      name_vi_moi | name_vi_new | ten_tieng_viet_moi
      name_en_moi | name_en_new | ten_tieng_anh_moi
  - Tuỳ chọn: tên cũ để kiểm tra trước khi ghi (so khớp không phân biệt hoa thường):
      name_vi_cu | name_vi_old | ten_tieng_viet_cu

Ví dụ:
  python3 update_analysis_item_names.py --csv ../csv/doi_ten_chi_tieu.csv --dry-run
  python3 update_analysis_item_names.py --xlsx ../../../data/Capability.xlsx --dry-run
  python3 update_analysis_item_names.py --xlsx  # dùng đường dẫn Capability mặc định trong repo

Chuỗi kết nối: biến môi trường VIETLABS_ODBC; nếu trống thì dùng cùng mẫu với các script import khác trong thư mục này.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import pyodbc
except ImportError:
    pyodbc = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


def _nh(s: str) -> str:
    """Chuẩn hoá header: NFKC, lower, gộp khoảng trắng."""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s)).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def _letters_all_upper(s: str) -> bool:
    letters = [c for c in s if c.isalpha()]
    return bool(letters) and all(c.isupper() and not c.islower() for c in letters)


def _letters_all_lower(s: str) -> bool:
    letters = [c for c in s if c.isalpha()]
    return bool(letters) and all(c.islower() for c in letters)


def _preserve_charwise_with_suffix(ref: str, target: str) -> str:
    """Áp hoa/thường theo vị trí ký tự; phần target dài hơn dùng tỷ lệ HOA trong ref."""
    letters = [c for c in ref if c.isalpha()]
    if letters:
        upper_share = sum(1 for c in letters if c.isupper() and not c.islower()) / len(letters)
    else:
        upper_share = 0.5

    out: List[str] = []
    for i, nc in enumerate(target):
        if i < len(ref):
            rc = ref[i]
            if rc.isalpha() and nc.isalpha():
                if rc.isupper() and not rc.islower():
                    out.append(nc.upper())
                elif rc.islower():
                    out.append(nc.lower())
                else:
                    out.append(nc)
            else:
                out.append(nc)
        else:
            rest = target[i:]
            if upper_share > 0.65:
                out.append(rest.upper())
            elif upper_share < 0.35:
                out.append(rest.lower())
            else:
                out.append(rest)
            break
    return "".join(out)


def preserve_case_from_reference(reference: str, target: str) -> str:
    """
    Áp pattern hoa/thường của `reference` (tên cũ) lên nội dung `target` (tên mới).

    - Cả chuỗi reference là HOA / thường (Unicode): áp upper()/lower() cho cả target.
    - Nhiều từ và cùng số từ với target: xử lý từng từ — từ reference toàn chữ HOA
      (vd. H2O) thì từ target tương ứng được upper (vd. nước → NƯỚC), tránh lệch vị trí số/chữ.
    - Còn lại: theo từng cặp ký tự cùng vị trí + quy tắc suffix như `_preserve_charwise_with_suffix`.
    """
    if not target:
        return target
    if not reference:
        return target

    ref = reference
    if ref.isupper():
        return target.upper()
    if ref.islower():
        return target.lower()

    ref_words = ref.split()
    tgt_words = target.split()
    if len(ref_words) == len(tgt_words) and len(ref_words) > 1:
        parts: List[str] = []
        for rw, tw in zip(ref_words, tgt_words):
            if _letters_all_upper(rw):
                parts.append(tw.upper())
            elif _letters_all_lower(rw):
                parts.append(tw.lower())
            else:
                parts.append(_preserve_charwise_with_suffix(rw, tw))
        return " ".join(parts)

    return _preserve_charwise_with_suffix(ref, target)


def _norm_name_compare(a: str, b: str) -> bool:
    """So sánh tên để xác minh (không phân biệt hoa thường, trim)."""
    return (a or "").strip().casefold() == (b or "").strip().casefold()


# --- alias cột sau khi _nh() ---
_CODE_KEYS = ("analysis_item_code", "mã chỉ tiêu", "ma chi tieu", "ma_chi_tieu")
_VI_NEW_KEYS = ("name_vi_moi", "name_vi_new", "ten_tieng_viet_moi", "ten_chi_tieu_moi")
_EN_NEW_KEYS = ("name_en_moi", "name_en_new", "ten_tieng_anh_moi")
_VI_OLD_KEYS = ("name_vi_cu", "name_vi_old", "ten_tieng_viet_cu", "ten_chi_tieu_cu")


def _pick(row_norm: Dict[str, str], keys: Tuple[str, ...]) -> str:
    for k in keys:
        nk = _nh(k)
        for rk, rv in row_norm.items():
            if rk == nk and rv.strip():
                return rv.strip()
    return ""


def _row_to_norm_keys(raw: Dict[str, str]) -> Dict[str, str]:
    return {_nh(k): (v if v is not None else "").strip() for k, v in raw.items()}


def _find_code_column(fieldnames: List[str]) -> Optional[str]:
    for fn in fieldnames:
        n = _nh(fn)
        if n in {_nh(x) for x in _CODE_KEYS}:
            return fn
        if "mã" in n and "chỉ" in n and "tiêu" in n:
            return fn
        if n.replace(" ", "") == "machitieu":
            return fn
    return None


def _resolve_column(fieldnames: List[str], aliases: Tuple[str, ...]) -> Optional[str]:
    alias_set = {_nh(a) for a in aliases}
    for fn in fieldnames:
        if _nh(fn) in alias_set:
            return fn
    return None


DEFAULT_CONNECTION = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=192.168.2.21,49938;"
    "Database=VietLabs;"
    "UID=sa;"
    "PWD=Sbt@2024;"
    "TrustServerCertificate=yes;"
    "Encrypt=yes;"
    "Login Timeout=60;"
)

# Cùng layout repo với link_analysis_item_analysis_group.py
DEFAULT_CAPABILITY_XLSX = str(Path(__file__).resolve().parents[3] / "data" / "Capability.xlsx")


def load_csv_rows(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV không có header.")
        fieldnames = list(reader.fieldnames)
        rows = [dict(r) for r in reader]
    return fieldnames, rows


def fetch_item_by_code(cursor, code: str) -> Optional[Tuple[str, str, str]]:
    cursor.execute(
        """
        SELECT analysis_item_id, name_vi, name_en
        FROM analysis_item
        WHERE analysis_item_code = ?
        """,
        (code.strip(),),
    )
    r = cursor.fetchone()
    if not r:
        return None
    return str(r[0]), (r[1] or "").strip(), (r[2] or "").strip()


def _norm_header_xlsx(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace('"', "").replace("'", "").strip()
    return re.sub(r"\s+", " ", s).lower()


def _xlsx_cell_text(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _find_code_name_col_indices(header_row: Tuple[Any, ...]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    pairs: List[Tuple[str, int]] = [(_norm_header_xlsx(c), i) for i, c in enumerate(header_row)]

    def find_i(pred) -> Optional[int]:
        for nh, idx in pairs:
            if nh and pred(nh):
                return idx
        return None

    code_i = find_i(lambda nh: "mã chỉ tiêu" in nh)
    nv_i = find_i(lambda nh: nh == "tên chỉ tiêu")
    ne_i = find_i(lambda nh: nh == "tên tiếng anh")
    return code_i, nv_i, ne_i


def read_capability_xlsx_name_rows(
    xlsx_path: str,
    sheet_names: List[str],
) -> Tuple[List[Tuple[str, str, str]], List[str]]:
    """
    Đọc các sheet, gộp theo mã: lần đầu gặp mã giữ (sheet đứng trước trong sheet_names được ưu tiên).
    Trả về danh sách (code, tên_vi_excel, tên_en_excel) và log trùng mã khác nội dung.
    """
    if openpyxl is None:
        raise RuntimeError("Cần cài openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    merged: Dict[str, Tuple[str, str]] = {}
    dup_notes: List[str] = []

    try:
        for sheet_name in sheet_names:
            name = sheet_name.strip()
            if not name:
                continue
            if name not in wb.sheetnames:
                dup_notes.append(f"[sheet] Bỏ qua sheet không tồn tại: {name!r} (có: {wb.sheetnames})")
                continue
            ws = wb[name]
            rows_it: Iterable[Tuple[Any, ...]] = ws.iter_rows(values_only=True)
            header = next(rows_it, None)
            if not header:
                dup_notes.append(f"[sheet] Sheet {name!r} không có header.")
                continue
            code_i, nv_i, ne_i = _find_code_name_col_indices(tuple(header))
            if code_i is None:
                dup_notes.append(f"[sheet] Sheet {name!r}: không tìm thấy cột 'Mã chỉ tiêu'.")
                continue
            if nv_i is None and ne_i is None:
                dup_notes.append(f"[sheet] Sheet {name!r}: không có cột Tên chỉ tiêu / Tên tiếng anh.")
                continue

            for ridx, row in enumerate(rows_it, start=2):
                if not row:
                    continue
                code = _xlsx_cell_text(row[code_i] if code_i < len(row) else None)
                if not code:
                    continue
                vi = _xlsx_cell_text(row[nv_i] if nv_i is not None and nv_i < len(row) else None)
                en = _xlsx_cell_text(row[ne_i] if ne_i is not None and ne_i < len(row) else None)
                if not vi and not en:
                    continue
                if code not in merged:
                    merged[code] = (vi, en)
                    continue
                pv, pe = merged[code]
                if (vi, en) == (pv, pe):
                    continue
                dup_notes.append(
                    f"{code} sheet={name!r} dòng={ridx}: khác bản đã chọn (giữ sheet trước trong --sheets); "
                    f"đã chọn VI={pv!r} EN={pe!r}, bỏ qua VI={vi!r} EN={en!r}"
                )
    finally:
        wb.close()

    ordered = list(merged.items())
    out = [(c, v, e) for c, (v, e) in ordered]
    return out, dup_notes


def main() -> int:
    if pyodbc is None:
        print("Cần cài pyodbc: pip install pyodbc", file=sys.stderr)
        return 1

    p = argparse.ArgumentParser(description="Cập nhật tên chỉ tiêu, giữ pattern hoa/thường.")
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv", metavar="PATH", help="File CSV nguồn tên mới")
    src.add_argument(
        "--xlsx",
        nargs="?",
        const="__default__",
        metavar="PATH",
        help="Capability.xlsx; không kèm PATH → dùng data/Capability.xlsx trong repo",
    )
    p.add_argument(
        "--sheets",
        default="Vietlabs,NTP",
        help="Với --xlsx: các sheet (phân cách bằng dấu phẩy). Mã trùng giữ sheet đứng trước.",
    )
    p.add_argument("--dry-run", action="store_true", help="Chỉ in preview, không ghi DB")
    args = p.parse_args()

    source_label = ""
    xlsx_notes: List[str] = []
    # (code, vi_new_src, en_new_src, vi_old_expected_or_none, preserve_case_from_db)
    flat_rows: List[Tuple[str, str, str, str, bool]] = []

    if args.csv is not None:
        csv_path = os.path.abspath(os.path.expanduser(args.csv))
        if not os.path.isfile(csv_path):
            here = os.path.dirname(os.path.abspath(__file__))
            sample = os.path.normpath(os.path.join(here, "..", "doi_ten_chi_tieu.csv.example"))
            print(f"Không tìm thấy file CSV: {csv_path}", file=sys.stderr)
            print(
                "Tạo file UTF-8 (có thể copy từ mẫu): analysis_item_code + name_vi_moi và/hoặc name_en_moi.",
                file=sys.stderr,
            )
            if os.path.isfile(sample):
                print(f"Mẫu trong repo: {sample}", file=sys.stderr)
                print(f"  cp {sample} doi_ten_chi_tieu.csv  # rồi sửa nội dung", file=sys.stderr)
            return 1

        fieldnames, raw_rows = load_csv_rows(csv_path)
        code_col = _find_code_column(fieldnames)
        if not code_col:
            print("Không tìm thấy cột mã chỉ tiêu (analysis_item_code / Mã chỉ tiêu).", file=sys.stderr)
            return 1

        col_vi_new = _resolve_column(fieldnames, _VI_NEW_KEYS)
        col_en_new = _resolve_column(fieldnames, _EN_NEW_KEYS)
        col_vi_old = _resolve_column(fieldnames, _VI_OLD_KEYS)

        if not col_vi_new and not col_en_new:
            print(
                "Cần ít nhất một cột tên mới: name_vi_moi / name_vi_new hoặc name_en_moi / name_en_new.",
                file=sys.stderr,
            )
            return 1

        for raw in raw_rows:
            rn = _row_to_norm_keys(raw)
            code = (raw.get(code_col) or "").strip()
            if not code:
                continue
            vi_new_src = _pick(rn, _VI_NEW_KEYS) if col_vi_new else ""
            en_new_src = _pick(rn, _EN_NEW_KEYS) if col_en_new else ""
            vi_old_expected = _pick(rn, _VI_OLD_KEYS) if col_vi_old else ""
            if not vi_new_src and not en_new_src:
                continue
            flat_rows.append((code, vi_new_src, en_new_src, vi_old_expected, True))

        source_label = f"CSV {csv_path}"

    else:
        if openpyxl is None:
            print("Chế độ --xlsx cần openpyxl: pip install openpyxl", file=sys.stderr)
            return 1
        if args.xlsx == "__default__":
            xlsx_path = DEFAULT_CAPABILITY_XLSX
        else:
            xlsx_path = os.path.abspath(os.path.expanduser(args.xlsx or ""))
        if not xlsx_path or not os.path.isfile(xlsx_path):
            print(f"Không tìm thấy file Excel: {xlsx_path!r}", file=sys.stderr)
            return 1

        sheet_list = [s.strip() for s in (args.sheets or "").split(",") if s.strip()]
        if not sheet_list:
            print("--sheets không được rỗng.", file=sys.stderr)
            return 1

        try:
            merged, xlsx_notes = read_capability_xlsx_name_rows(xlsx_path, sheet_list)
        except Exception as e:
            print(f"Lỗi đọc Excel: {e}", file=sys.stderr)
            return 1

        for code, vi, en in merged:
            flat_rows.append((code, vi, en, "", False))

        source_label = f"Excel {xlsx_path} sheets={','.join(sheet_list)}"

    cs = (os.environ.get("VIETLABS_ODBC") or "").strip() or DEFAULT_CONNECTION

    conn = pyodbc.connect(cs)
    conn.autocommit = False
    cur = conn.cursor()

    updates: List[Tuple[str, str, Optional[str], str, Optional[str], str, Optional[str]]] = []
    # (analysis_item_id, code, old_vi, new_vi_sql, old_en, new_en_sql, note)
    skipped: List[str] = []

    for code, vi_new_src, en_new_src, vi_old_expected, preserve_case_from_db in flat_rows:
        got = fetch_item_by_code(cur, code)
        if not got:
            skipped.append(f"{code}: không tìm thấy analysis_item (theo mã)")
            continue
        aid, db_vi, db_en = got

        if vi_old_expected and not _norm_name_compare(vi_old_expected, db_vi):
            skipped.append(
                f"{code}: tên VI trong DB không khớp cột tên cũ (DB={db_vi!r}, file={vi_old_expected!r})"
            )
            continue

        new_vi: Optional[str] = None
        new_en: Optional[str] = None
        if vi_new_src:
            if preserve_case_from_db:
                new_vi = preserve_case_from_reference(db_vi, vi_new_src) if db_vi else vi_new_src
            else:
                new_vi = vi_new_src
        if en_new_src:
            if preserve_case_from_db:
                new_en = preserve_case_from_reference(db_en, en_new_src) if db_en else en_new_src
            else:
                new_en = en_new_src

        if new_vi == db_vi:
            new_vi = None
        if new_en == db_en:
            new_en = None
        if new_vi is None and new_en is None:
            continue

        updates.append((aid, code, db_vi, new_vi, db_en, new_en, None))

    for note in xlsx_notes[:50]:
        print(f"  [XLSX] {note}")
    if len(xlsx_notes) > 50:
        print(f"  ... và {len(xlsx_notes) - 50} dòng ghi chú XLSX khác")

    print(
        f"Nguồn: {source_label}. Dòng có mã + tên: {len(flat_rows)}. "
        f"Sẽ cập nhật: {len(updates)}, bỏ qua / lỗi: {len(skipped)}, dry_run={args.dry_run}"
    )
    for s in skipped[:40]:
        print(f"  [SKIP] {s}")
    if len(skipped) > 40:
        print(f"  ... và {len(skipped) - 40} dòng skip khác")

    for u in updates[:25]:
        aid, code, old_vi, nv, old_en, ne, _ = u
        if nv is not None:
            print(f"  {code}: name_vi {old_vi!r} -> {nv!r}")
        if ne is not None:
            print(f"  {code}: name_en {old_en!r} -> {ne!r}")
    if len(updates) > 25:
        print(f"  ... và {len(updates) - 25} dòng cập nhật khác")

    if args.dry_run:
        conn.close()
        return 0

    now = datetime.now(timezone.utc)
    try:
        # Chỉ name_vi / name_en (+ updated_at). Không SET analysis_item_code.
        for aid, code, old_vi, nv, old_en, ne, _ in updates:
            if nv is not None and ne is not None:
                cur.execute(
                    """
                    UPDATE analysis_item
                    SET name_vi = ?, name_en = ?, updated_at = ?
                    WHERE analysis_item_id = ?
                    """,
                    (nv, ne, now, aid),
                )
            elif nv is not None:
                cur.execute(
                    "UPDATE analysis_item SET name_vi = ?, updated_at = ? WHERE analysis_item_id = ?",
                    (nv, now, aid),
                )
            elif ne is not None:
                cur.execute(
                    "UPDATE analysis_item SET name_en = ?, updated_at = ? WHERE analysis_item_id = ?",
                    (ne, now, aid),
                )
        conn.commit()
        print(f"Đã commit {len(updates)} bản ghi.")
    except Exception as e:
        conn.rollback()
        print(f"Lỗi, rollback: {e}", file=sys.stderr)
        return 1
    finally:
        conn.close()
    return 0


def _self_test() -> None:
    assert preserve_case_from_reference("PHÂN TÍCH H2O", "Phân tích nước") == "PHÂN TÍCH NƯỚC"
    assert preserve_case_from_reference("Phân tích H2O", "phân tích nước") == "Phân tích NƯỚC"
    assert preserve_case_from_reference("pH đo", "ph đo") == "pH đo"
    assert preserve_case_from_reference("ABC", "xyz") == "XYZ"


if __name__ == "__main__":
    if len(sys.argv) == 2 and sys.argv[1] == "--self-test":
        _self_test()
        print("preserve_case_from_reference: OK")
        sys.exit(0)
    sys.exit(main())
