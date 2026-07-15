#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Regenerate Capability*SeedData.cs và cập nhật CapabilityImportRules sample matrix groups
từ workbook Danh mục Năng lực v2.

  python3 regenerate_capability_seed_data_from_xlsx.py
  python3 regenerate_capability_seed_data_from_xlsx.py --xlsx path --write
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent))
import import_analysis_item as iai

try:
    import openpyxl
except ImportError:
    print("Cần: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_REPO = Path(__file__).resolve().parents[3]
_DATA = _REPO / "data"
_DEFAULT_GLOB = list(_DATA.glob("*N*ng l*c v2.xlsx"))
DEFAULT_XLSX = _DEFAULT_GLOB[0] if _DEFAULT_GLOB else _DATA / "Danh mục Năng lực v2.xlsx"
_OUT_DIR = _REPO / "Vietlabs-API" / "Data"

# English names for sample matrix groups (Layer0 seeder)
_SMO_EN: Dict[str, str] = {
    "Bao bì tiếp xúc với thực phẩm": "Food contact packaging",
    "Bao bì tiếp xúc với Thực phẩm": "Food contact packaging",
    "Dầu mỡ động thực vật": "Animal and vegetable oils and fats",
    "Hóa chất": "Chemicals",
    "Mẫu bệnh phẩm thú y": "Veterinary specimens",
    "Mẫu vệ sinh bề mặt": "Surface hygiene samples",
    "Mỹ phẩm": "Cosmetics",
    "Nước": "Water",
    "Nước uống có cồn và không cồn": "Alcoholic and non-alcoholic beverages",
    "Phân bón, chế phẩm sinh học": "Fertilizers and biological products",
    "Phụ gia thực phẩm": "Food additives",
    "phụ gia Thực phẩm": "Food additives",
    "Thức ăn và nguyên liệu thức ăn": "Animal feed and feed ingredients",
    "Thực phẩm": "Food",
    "Đất": "Soil",
}


def norm_header(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).replace("\n", " ").replace('"', "").strip()).lower()


def norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip()).casefold()


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def cs_string(s: str) -> str:
    s = re.sub(r"\s+", " ", s.replace("\r\n", "\n").replace("\r", "\n").strip())
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def read_workbook(path: Path) -> Tuple[Dict[str, List[Tuple[Any, ...]]], Dict[str, Dict[str, int]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: Dict[str, List[Tuple[Any, ...]]] = {}
    maps: Dict[str, Dict[str, int]] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = tuple(rows[0])
        pairs = [(norm_header(v), i) for i, v in enumerate(header)]

        def find(pred) -> Optional[int]:
            for nh, idx in pairs:
                if nh and pred(nh):
                    return idx
            return None

        m: Dict[str, int] = {}
        for k, pred in [
            ("sm_group", lambda nh: nh == "nhóm nền mẫu"),
            ("sm", lambda nh: nh == "nền mẫu"),
            ("standard", lambda nh: "tiêu chuẩn" in nh and ("qui chuẩn" in nh or "quy chuẩn" in nh)),
            ("analysis_group", lambda nh: nh == "nhóm chỉ tiêu"),
            ("reference_method", lambda nh: nh == "phương pháp"),
            ("equipment", lambda nh: "thiết bị" in nh or "equipment" in nh),
            ("registered", lambda nh: "nền mẫu đã đăng ký" in nh),
            ("group_price", lambda nh: "giá nhóm chuẩn_new" in nh or nh == "giá nhóm chuẩn"),
        ]:
            idx = find(pred)
            if idx is not None:
                m[k] = idx
        sheets[sn] = rows[1:]
        maps[sn] = m
    wb.close()
    return sheets, maps


def collect_distinct(
    sheets: Dict[str, List[Tuple[Any, ...]]],
    maps: Dict[str, Dict[str, int]],
) -> Dict[str, Any]:
    standards: Set[str] = set()
    ref_methods: Set[str] = set()
    equipment: Set[str] = set()
    sm_groups: Dict[str, str] = {}  # norm -> canonical
    sm_pairs: Dict[Tuple[str, str], str] = {}  # (group, matrix) -> registered sample
    ag_prices: Dict[str, List[Decimal]] = defaultdict(list)

    for sn, rows in sheets.items():
        m = maps[sn]
        for row in rows:
            if not row:
                continue

            def g(key: str) -> str:
                idx = m.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return cell_str(row[idx])

            std = g("standard")
            if std and std.upper() != "NA":
                standards.add(std)
            rm = g("reference_method")
            if rm:
                ref_methods.add(rm)
            eq = g("equipment")
            if eq and eq.upper() != "NA":
                equipment.add(eq)
            sg = g("sm_group")
            sm = g("sm")
            reg = g("registered")
            if sg:
                sm_groups[norm_key(sg)] = sg
            if sg and sm:
                k = (norm_key(sg), norm_key(sm))
                if k not in sm_pairs or (reg and not sm_pairs[k]):
                    sm_pairs[k] = reg or sm
            ag = g("analysis_group")
            gp = g("group_price")
            if ag and not iai.is_blank_analysis_group_cell(ag):
                if gp and gp.upper() != "NA":
                    try:
                        ag_prices[ag].append(Decimal(str(gp).replace(",", "")))
                    except (InvalidOperation, ValueError):
                        pass

    # analysis group whole price = mode per group
    ag_rows: List[Tuple[str, Optional[Decimal]]] = []
    for name in sorted(ag_prices.keys(), key=lambda x: (x.upper(), x)):
        vals = ag_prices[name]
        if not vals:
            ag_rows.append((name, None))
            continue
        freq: Dict[Decimal, int] = defaultdict(int)
        for v in vals:
            freq[v] += 1
        mode_val = max(freq.items(), key=lambda x: (x[1], x[0]))[0]
        ag_rows.append((name, mode_val))

    sm_group_list = sorted(set(sm_groups.values()), key=lambda x: (x.upper(), x))
    sm_rows = []
    seen_pair: Set[Tuple[str, str]] = set()
    for (gk, mk), reg in sorted(sm_pairs.items(), key=lambda x: (x[0][0], x[0][1])):
        # recover canonical names from sm_groups keys - use first matching
        gv = next(v for k, v in sm_groups.items() if k == gk)
        # matrix canonical from rows - find original
        mv = None
        for sn, rows in sheets.items():
            m = maps[sn]
            for row in rows:
                sg = cell_str(row[m["sm_group"]]) if "sm_group" in m and m["sm_group"] < len(row) else ""
                sm = cell_str(row[m["sm"]]) if "sm" in m and m["sm"] < len(row) else ""
                if norm_key(sg) == gk and norm_key(sm) == mk:
                    mv = sm
                    break
            if mv:
                break
        if not mv:
            continue
        if (gk, mk) in seen_pair:
            continue
        seen_pair.add((gk, mk))
        sm_rows.append((gv, mv, reg or mv))

    return {
        "standards": sorted(standards, key=lambda x: (x.upper(), x)),
        "reference_methods": sorted(ref_methods, key=lambda x: (x.upper(), x)),
        "equipment": sorted(equipment, key=lambda x: (x.upper(), x)),
        "sm_groups": sm_group_list,
        "sm_rows": sm_rows,
        "analysis_groups": ag_rows,
    }


def write_standards(values: List[str]) -> None:
    lines = [
        "namespace VietLab.Data;",
        "",
        "/// <summary>",
        "/// Giá trị distinct cột Quy chuẩn/Tiêu chuẩn từ sheet Vietlabs + NTP (Danh mục Năng lực v2).",
        "/// Sinh tự động — cập nhật khi đổi Excel.",
        "/// </summary>",
        "internal static class CapabilityStandardsSeedData",
        "{",
        "    public static readonly string[] NameViList =",
        "    [",
    ]
    for v in values:
        lines.append(f"        {cs_string(v)},")
    lines.extend(["    ];", "}", ""])
    (_OUT_DIR / "CapabilityStandardsSeedData.cs").write_text("\n".join(lines), encoding="utf-8")


def write_string_list(class_name: str, field_name: str, comment: str, values: List[str]) -> None:
    lines = [
        "namespace VietLab.Data;",
        "",
        "/// <summary>",
        f"/// {comment}",
        "/// Sinh tự động — cập nhật khi đổi Excel.",
        "/// </summary>",
        f"internal static class {class_name}",
        "{",
        f"    public static readonly string[] {field_name} =",
        "    [",
    ]
    for v in values:
        lines.append(f"        {cs_string(v)},")
    lines.extend(["    ];", "}", ""])
    (_OUT_DIR / f"{class_name}.cs").write_text("\n".join(lines), encoding="utf-8")


def write_sample_matrices(rows: List[Tuple[str, str, str]]) -> None:
    lines = [
        "namespace VietLab.Data;",
        "",
        "",
        "/// <summary>",
        "/// Dữ liệu seed nền mẫu: distinct (Nhóm nền mẫu, Nền mẫu) từ sheet Vietlabs + NTP (Danh mục Năng lực v2).",
        "/// Sinh tự động — cập nhật bằng script khi đổi Excel.",
        "/// </summary>",
        "internal static class CapabilitySampleMatricesSeedData",
        "{",
        "    public static readonly SampleMatrixSeed[] Rows =",
        "    [",
    ]
    for gv, mv, reg in rows:
        lines.append(f"        new({cs_string(gv)}, {cs_string(mv)}, {cs_string(reg)}),")
    lines.extend(["    ];", "}", ""])
    (_OUT_DIR / "CapabilitySampleMatricesSeedData.cs").write_text("\n".join(lines), encoding="utf-8")


def write_analysis_groups(rows: List[Tuple[str, Optional[Decimal]]]) -> None:
    lines = [
        "namespace VietLab.Data;",
        "",
        "/// <summary>",
        "/// Nhóm chỉ tiêu distinct từ Danh mục Năng lực v2 sheet Vietlabs:",
        "/// cột Nhóm Chỉ tiêu + Giá nhóm chuẩn_new (mode theo nhóm).",
        "/// </summary>",
        "internal static class CapabilityAnalysisGroupsSeedData",
        "{",
        "    public readonly record struct Row(string NameVi, string NameEn, decimal? WholeGroupStandardPrice);",
        "",
        "    public static readonly Row[] Rows =",
        "    [",
    ]
    for name, price in rows:
        if iai.is_blank_analysis_group_cell(name):
            continue
        price_str = "null" if price is None else f"{price}m"
        lines.append(f"        new({cs_string(name)}, {cs_string(name)}, {price_str}),")
    lines.extend(["    ];", "}", ""])
    (_OUT_DIR / "CapabilityAnalysisGroupsSeedData.cs").write_text("\n".join(lines), encoding="utf-8")


def patch_import_rules_sm_groups(groups: List[str]) -> None:
    path = _OUT_DIR / "CapabilityImportRules.cs"
    text = path.read_text(encoding="utf-8")
    entries = []
    for g in groups:
        en = _SMO_EN.get(g, g)
        entries.append(f'            new("{g}", "{en}"),')
    block = "\n".join(entries)
    new_block = f"""    public static readonly ReadOnlyCollection<SampleMatrixGroupSeed> SampleMatrixGroupsFromCapabilityXlsx =
        new(
        [
{block}
        ]);"""
    text = re.sub(
        r"public static readonly ReadOnlyCollection<SampleMatrixGroupSeed> SampleMatrixGroupsFromCapabilityXlsx =[\s\S]*?\]\);",
        new_block,
        text,
        count=1,
    )
    path.write_text(text, encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--write", action="store_true", help="Ghi file C# (mặc định chỉ in thống kê)")
    args = ap.parse_args()
    if not args.xlsx.is_file():
        print(f"Không tìm thấy: {args.xlsx}", file=sys.stderr)
        return 2

    sheets, maps = read_workbook(args.xlsx)
    data = collect_distinct(sheets, maps)
    print(f"standards: {len(data['standards'])}")
    print(f"reference_methods: {len(data['reference_methods'])}")
    print(f"equipment: {len(data['equipment'])}")
    print(f"sm_groups: {len(data['sm_groups'])}")
    print(f"sm_rows: {len(data['sm_rows'])}")
    print(f"analysis_groups: {len(data['analysis_groups'])}")

    if not args.write:
        print("Chạy với --write để ghi file seed.")
        return 0

    write_standards(data["standards"])
    write_string_list(
        "CapabilityReferenceMethodsSeedData",
        "NameViList",
        'Giá trị distinct cột "Phương pháp" từ sheet Vietlabs + NTP (Danh mục Năng lực v2).',
        data["reference_methods"],
    )
    write_string_list(
        "CapabilityEquipmentTypesSeedData",
        "NameViList",
        'Giá trị distinct cột "Thiết bị/ Equipment" từ sheet Vietlabs + NTP (Danh mục Năng lực v2).',
        data["equipment"],
    )
    write_sample_matrices(data["sm_rows"])
    write_analysis_groups(data["analysis_groups"])
    patch_import_rules_sm_groups(data["sm_groups"])
    print(f"Đã ghi seed vào {_OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
