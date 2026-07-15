#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Chuẩn hóa sheet NTP bổ sung: mã CT, giá, sentinel, nền mẫu thiếu."""
from __future__ import annotations

import argparse
import csv
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Set, Tuple

try:
    import openpyxl
except ImportError:
    raise SystemExit("Can cai: pip install openpyxl")

SHEET = "NTP bổ sung"
START_CODE_NUM = 11219
_RE_CT = re.compile(r"(?i)^CT-(\d+)$")


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def is_blank_sentinel(s: str) -> bool:
    if not s or not str(s).strip():
        return True
    u = str(s).strip().upper().replace(" ", "")
    return u in ("", "NA", "N/A", "-", "--", "NONE", "NULL", "KHÔNG", "KHONGCO")


def norm_key(name: str, method: str, sm: str, sm_group: str) -> Tuple[str, str, str, str]:
    def n(x: str) -> str:
        return re.sub(r"\s+", " ", (x or "").strip().casefold())

    return n(name), n(method), n(sm), n(sm_group)


def collect_existing_codes(wb) -> Set[str]:
    codes: Set[str] = set()
    for sn in wb.sheetnames:
        if sn == SHEET:
            continue
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        code_i = next(
            (i for i, h in enumerate(header) if h and "mã chỉ tiêu" in str(h).lower()),
            None,
        )
        if code_i is None:
            continue
        for row in rows:
            if code_i < len(row):
                c = cell_str(row[code_i])
                if _RE_CT.match(c):
                    codes.add(c.upper())
    return codes


def find_header_indices(header: Tuple[Any, ...]) -> dict:
    idx = {}
    for i, h in enumerate(header):
        nh = cell_str(h).lower()
        if "mã chỉ tiêu" in nh:
            idx["code"] = i
        elif nh == "nhóm nền mẫu":
            idx["sm_group"] = i
        elif nh == "nền mẫu":
            idx["sm"] = i
        elif nh == "tên chỉ tiêu":
            idx["name_vi"] = i
        elif nh == "phương pháp":
            idx["reference_method"] = i
        elif "đơn giá" in nh and "ntp" in nh:
            idx["unit_price"] = i
    return idx


def normalize_price_value(v: Any) -> Optional[Any]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = float(v)
        if n <= 0:
            return None
        if 1 <= n <= 9999 and n == int(n):
            return int(n * 1000)
        return int(n) if n == int(n) else n
    s = cell_str(v)
    if is_blank_sentinel(s):
        return None
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    return int(digits)


def prepare_workbook(xlsx_path: Path, dry_run: bool = False) -> Path:
    if not xlsx_path.is_file():
        raise SystemExit(f"Khong tim thay: {xlsx_path}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = xlsx_path.with_name(f"{xlsx_path.stem}_backup_{ts}{xlsx_path.suffix}")
    if not dry_run:
        shutil.copy2(xlsx_path, backup)
        print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(xlsx_path)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Thieu sheet {SHEET!r}")

    existing_codes = collect_existing_codes(wb)
    max_num = 0
    for c in existing_codes:
        m = _RE_CT.match(c)
        if m:
            max_num = max(max_num, int(m.group(1)))

    start_num = max(START_CODE_NUM, max_num + 1)
    ws = wb[SHEET]
    header = tuple(ws.iter_rows(min_row=1, max_row=1, values_only=True)).__iter__().__next__()
    cols = find_header_indices(header)

    required = ("code", "sm_group", "sm", "name_vi", "unit_price")
    missing = [k for k in required if k not in cols]
    if missing:
        raise SystemExit(f"Thieu cot: {missing}")

    warnings: List[Tuple[int, str, str]] = []
    assigned_codes: List[str] = []
    code_num = start_num

    for ridx in range(2, ws.max_row + 1):
        name_vi = cell_str(ws.cell(ridx, cols["name_vi"] + 1).value)
        if not name_vi:
            continue

        new_code = f"CT-{code_num:05d}"
        if new_code.upper() in existing_codes:
            raise SystemExit(f"Trung ma {new_code}")
        assigned_codes.append(new_code)
        existing_codes.add(new_code.upper())
        code_num += 1

        if not dry_run:
            ws.cell(ridx, cols["code"] + 1).value = new_code

        sm_val = cell_str(ws.cell(ridx, cols["sm"] + 1).value)
        sm_group_val = cell_str(ws.cell(ridx, cols["sm_group"] + 1).value)
        if not sm_val and sm_group_val:
            if not dry_run:
                ws.cell(ridx, cols["sm"] + 1).value = sm_group_val
            warnings.append((ridx, new_code, f"Bo sung Nền mẫu = {sm_group_val!r}"))

        price_cell = ws.cell(ridx, cols["unit_price"] + 1)
        old_price = price_cell.value
        new_price = normalize_price_value(old_price)
        if not dry_run:
            price_cell.value = new_price

        ref = cell_str(ws.cell(ridx, cols.get("reference_method", -1) + 1).value) if "reference_method" in cols else ""
        key = norm_key(name_vi, ref, sm_val or sm_group_val, sm_group_val)
        for sn in wb.sheetnames:
            if sn == SHEET:
                continue
            ows = wb[sn]
            oheader = tuple(ows.iter_rows(min_row=1, max_row=1, values_only=True)).__iter__().__next__()
            ocols = find_header_indices(oheader)
            if "code" not in ocols:
                continue
            for orow in range(2, ows.max_row + 1):
                oname = cell_str(ows.cell(orow, ocols["name_vi"] + 1).value)
                if not oname:
                    continue
                oref = cell_str(ows.cell(orow, ocols.get("reference_method", -1) + 1).value) if "reference_method" in ocols else ""
                osm = cell_str(ows.cell(orow, ocols["sm"] + 1).value)
                osg = cell_str(ows.cell(orow, ocols["sm_group"] + 1).value)
                if norm_key(oname, oref, osm, osg) == key:
                    old_code = cell_str(ows.cell(orow, ocols["code"] + 1).value)
                    warnings.append((ridx, new_code, f"Gan giong ma cu {old_code} tren sheet {sn}"))

    if len(assigned_codes) != 160:
        print(f"Canh bao: gan {len(assigned_codes)} ma (khong phai 160)")

    report = xlsx_path.parent / f"{xlsx_path.stem}_ntp_bo_sung_prepare_{ts}.csv"
    if warnings:
        with open(report, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["sheet_row", "analysis_item_code", "reason"])
            w.writerows(warnings)
        print(f"Canh bao: {len(warnings)} dong -> {report}")

    if not dry_run:
        wb.save(xlsx_path)
        print(f"Da luu: {xlsx_path}")
    else:
        print("[dry-run] Khong ghi workbook")

    wb.close()
    print(f"Ma: {assigned_codes[0]} .. {assigned_codes[-1]} ({len(assigned_codes)} dong)")
    return backup if not dry_run else xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Chuan hoa sheet NTP bo sung")
    parser.add_argument(
        "--xlsx",
        default=str(Path(__file__).resolve().parents[3] / "data" / "Danh mục Năng lực v3.xlsx"),
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    prepare_workbook(Path(args.xlsx), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
