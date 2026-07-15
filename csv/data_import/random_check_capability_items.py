#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Random spot-check: so sánh N mã CT ngẫu nhiên giữa workbook v2 và DB.

  python3 random_check_capability_items.py --count 10
  python3 random_check_capability_items.py --count 20 --seed 42
  python3 random_check_capability_items.py --codes CT-0001,CT-0004

Kiểm tra mỗi mã:
  - analysis_item (name, display Tiptap, master FK text, TAT, giá)
  - department_analysis_capability + designation (Vietlabs)
  - subcontractor_capability + designation (NTP)
"""
from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import capability_workbook_paths as cwp
import excel_cell_tiptap as ect
import import_analysis_item as iai
import import_analysis_items_capability_vietlabs_xlsx as iac
import import_department_capability_vietlabs_xlsx as idc
import import_subcontractor_capability_ntp_xlsx as isc

try:
    import openpyxl
    import pyodbc
except ImportError:
    print("Cần: pip install openpyxl pyodbc", file=sys.stderr)
    sys.exit(1)

FORMAT_MARKS = frozenset({"italic", "subscript", "superscript"})


@dataclass
class ExcelRow:
    code: str
    sheet: str
    row_idx: int
    fields: Dict[str, Any] = field(default_factory=dict)


@dataclass
class CheckIssue:
    code: str
    area: str
    message: str


def norm_date(dt: Optional[datetime]) -> Optional[str]:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d")


def normalize_tiptap_runs(value: Optional[str]) -> List[Tuple[str, Tuple[str, ...]]]:
    """Chuẩn hóa Tiptap doc thành list (text, marks) để so sánh."""
    if not value or not str(value).strip():
        return []
    try:
        doc = json.loads(value) if isinstance(value, str) else value
    except json.JSONDecodeError:
        return [(str(value).strip(), ())]
    runs: List[Tuple[str, Tuple[str, ...]]] = []

    def walk(node: Any) -> None:
        if not isinstance(node, dict):
            return
        if node.get("type") == "text":
            marks = tuple(
                sorted(
                    m.get("type")
                    for m in (node.get("marks") or [])
                    if isinstance(m, dict) and m.get("type") in FORMAT_MARKS
                )
            )
            runs.append((node.get("text") or "", marks))
            return
        for child in node.get("content") or []:
            walk(child)

    walk(doc)
    return runs


def tiptap_equal(a: Optional[str], b: Optional[str]) -> bool:
    return normalize_tiptap_runs(a) == normalize_tiptap_runs(b)


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    return iai.normalize_text(str(value).strip())


def norm_price(value: Any) -> Optional[float]:
    if value is None:
        return None
    p = iai.parse_price(str(value))
    return round(float(p), 4) if p is not None else None


def index_workbook(xlsx: str) -> Dict[str, ExcelRow]:
    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=True, rich_text=True)
    out: Dict[str, ExcelRow] = {}
    for sheet in ("Vietlabs", "NTP"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = tuple(c.value for c in ws[1])
        colmap = iac.build_column_map(header)
        capmap, _ = idc.build_maps(header)
        if "code" not in colmap:
            continue

        for ridx in range(2, (ws.max_row or 1) + 1):
            code_cell = ws.cell(ridx, colmap["code"] + 1)
            code = iac.cell_str(code_cell.value)
            if not code or not iac.is_valid_ct_code(code):
                continue
            code = code.strip()

            def cell_val(key: str) -> Any:
                idx = colmap.get(key)
                if idx is None:
                    return None
                return ws.cell(ridx, idx + 1).value

            def cell_obj(key: str):
                idx = colmap.get(key)
                if idx is None:
                    return None
                return ws.cell(ridx, idx + 1)

            name_vi_cell = cell_obj("name_vi")
            name_en_cell = cell_obj("name_en")
            short_cell = cell_obj("short_name")

            name_vi_plain, display_vi = (
                ect.cell_to_plain_and_display(name_vi_cell) if name_vi_cell else (None, None)
            )
            name_en_plain, display_en = (
                ect.cell_to_plain_and_display(name_en_cell) if name_en_cell else (None, None)
            )
            short_plain, display_short = (
                ect.cell_to_plain_and_display(short_cell) if short_cell else (None, None)
            )

            hdr = tuple(header)
            tats: Dict[str, Optional[int]] = {}
            for tk in ("tat_normal", "tat_fast", "tat_urgent"):
                idx = colmap.get(tk)
                if idx is not None and idx < len(hdr):
                    hn = iac.norm_header_cell(hdr[idx])
                    tats[tk] = iac.parse_tat_cell_to_hours(
                        ws.cell(ridx, idx + 1).value, hn, sheet
                    )

            nd107: Dict[str, Tuple[bool, Optional[str]]] = {}
            for site, cap_key, branch in idc.SITES_ND107_FULL:
                idx = capmap.get(cap_key)
                if idx is None:
                    continue
                has, dt = idc.parse_nd107_cell(ws.cell(ridx, idx + 1).value)
                nd107[branch] = (has, norm_date(dt))

            designations: Dict[str, Dict[str, Optional[str]]] = {}
            for dk, dcode in idc.DESIGNATION_CODES.items():
                idx = capmap.get(dk)
                if idx is None:
                    continue
                exp = idc.parse_designation_cell(ws.cell(ridx, idx + 1).value)
                designations[dcode] = {"expired_date": norm_date(exp), "has_value": exp is not None}

            ntp_label = None
            ntp_nd107: Tuple[bool, Optional[str]] = (False, None)
            ntp_des: Dict[str, Optional[str]] = {}
            if sheet == "NTP":
                for nh, idx in ((idc.norm_header_cell(v), i) for i, v in enumerate(header)):
                    if "ghi chú" in nh and "tên ntp" in nh:
                        ntp_label = iac.cell_str(ws.cell(ridx, idx + 1).value) or None
                    if "nđ 107" in nh.replace(" ", "") or nh.startswith("nđ 107"):
                        reg = isc.parse_registered_cap_cell(ws.cell(ridx, idx + 1).value)
                        ntp_nd107 = (reg is not None, norm_date(reg.expired_date) if reg else None)
                for dk, dcode in isc.DESIGNATION_CODES.items():
                    idx = capmap.get(dk)
                    if idx is not None:
                        reg = isc.parse_registered_cap_cell(ws.cell(ridx, idx + 1).value)
                        if reg is not None:
                            ntp_des[dcode] = norm_date(reg.expired_date)

            fields: Dict[str, Any] = {
                "name_vi": name_vi_plain or iac.cell_str(name_vi_cell.value if name_vi_cell else ""),
                "display_name_vi": display_vi,
                "name_en": name_en_plain or iac.cell_str(name_en_cell.value if name_en_cell else ""),
                "display_name_en": display_en,
                "short_name": short_plain or iac.cell_str(short_cell.value if short_cell else "") or None,
                "display_short_name": display_short,
                "analysis_group": iac.cell_str(cell_val("analysis_group")),
                "standard": iac.cell_str(cell_val("standard")),
                "reference_method": iac.cell_str(cell_val("reference_method")),
                "uom": iac.cell_str(cell_val("uom")),
                "equipment": iac.cell_str(cell_val("equipment")),
                "sm_group": iac.cell_str(cell_val("sm_group")),
                "sm": iac.cell_str(cell_val("sm")),
                "lab_technique": iac.cell_str(cell_val("lab_technique")),
                "lod": iac.cell_str(cell_val("lod")),
                "loq": iac.cell_str(cell_val("loq")),
                "standard_value": iac.cell_str(cell_val("standard_value")),
                "unit_price": norm_price(cell_val("unit_price")),
                "status": iac.cell_str(cell_val("status")),
                "tats": tats,
                "nd107": nd107,
                "designations": designations,
                "ntp_label": ntp_label,
                "ntp_nd107": ntp_nd107,
                "ntp_designations": ntp_des,
            }
            out[code] = ExcelRow(code=code, sheet=sheet, row_idx=ridx, fields=fields)
    wb.close()
    return out


def fetch_db_snapshots(cur, codes: List[str]) -> Dict[str, Dict[str, Any]]:
    if not codes:
        return {}
    placeholders = ",".join("?" * len(codes))
    cur.execute(
        f"""
        SELECT
            ai.analysis_item_code,
            ai.name_vi, ai.name_en, ai.short_name,
            ai.display_name_vi, ai.display_name_en, ai.display_short_name,
            ai.lod, ai.loq, ai.standard_value, ai.unit_price, ai.status,
            ag.name_vi AS analysis_group,
            s.name_vi AS standard_name,
            rm.name_vi AS reference_method,
            uom.name_vi AS uom_name,
            et.name_vi AS equipment_name,
            smg.name_vi AS sm_group_name,
            sm.name_vi AS sm_name,
            lt.name_vi AS lab_technique
        FROM analysis_item ai
        LEFT JOIN analysis_group ag ON ai.analysis_group_id = ag.analysis_group_id
        LEFT JOIN standard s ON ai.standard_id = s.standard_id
        LEFT JOIN reference_method rm ON ai.reference_method_id = rm.reference_method_id
        LEFT JOIN unit_of_measure uom ON ai.unit_of_measure_id = uom.unit_of_measure_id
        LEFT JOIN equipment_type et ON ai.equipment_type_id = et.equipment_type_id
        LEFT JOIN sample_matrix sm ON ai.sample_matrix_id = sm.sample_matrix_id
        LEFT JOIN sample_matrix_group smg ON ai.sample_matrix_group_id = smg.sample_matrix_group_id
        LEFT JOIN laboratory_technique lt ON ai.laboratory_technique_id = lt.laboratory_technique_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    rows: Dict[str, Dict[str, Any]] = {}
    for r in cur.fetchall():
        code = str(r[0]).strip()
        rows[code] = {
            "name_vi": r[1],
            "name_en": r[2],
            "short_name": r[3],
            "display_name_vi": r[4],
            "display_name_en": r[5],
            "display_short_name": r[6],
            "lod": r[7],
            "loq": r[8],
            "standard_value": r[9],
            "unit_price": round(float(r[10]), 4) if r[10] is not None else None,
            "status": r[11],
            "analysis_group": r[12],
            "standard": r[13],
            "reference_method": r[14],
            "uom": r[15],
            "equipment": r[16],
            "sm_group": r[17],
            "sm": r[18],
            "lab_technique": r[19],
            "tats": {},
            "dac": {},
            "dac_des": {},
            "sc": [],
            "sc_des": {},
        }

    cur.execute(
        f"""
        SELECT ai.analysis_item_code, t.tat_type, t.tat_value
        FROM analysis_item_tat t
        JOIN analysis_item ai ON t.analysis_item_id = ai.analysis_item_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    tat_map = {"Normal": "tat_normal", "Fast": "tat_fast", "Urgent": "tat_urgent"}
    for code, tat_type, val in cur.fetchall():
        key = tat_map.get(str(tat_type))
        if key and code in rows:
            rows[code]["tats"][key] = int(val) if val is not None else None

    cur.execute(
        f"""
        SELECT ai.analysis_item_code, b.branch_code,
               dac.nd_107, dac.nd_107_expired_date
        FROM department_analysis_capability dac
        JOIN analysis_item ai ON dac.analysis_item_id = ai.analysis_item_id
        JOIN branch b ON dac.branch_id = b.branch_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    for code, branch, nd107, exp in cur.fetchall():
        if code not in rows:
            continue
        rows[code]["dac"][str(branch)] = {
            "nd_107": bool(nd107),
            "expired_date": norm_date(exp),
        }

    cur.execute(
        f"""
        SELECT ai.analysis_item_code, b.branch_code, d.designation_code, dacd.expired_date
        FROM department_analysis_capability_designation dacd
        JOIN department_analysis_capability dac
          ON dacd.department_analysis_capability_id = dac.department_analysis_capability_id
        JOIN analysis_item ai ON dac.analysis_item_id = ai.analysis_item_id
        JOIN branch b ON dac.branch_id = b.branch_id
        JOIN designation d ON dacd.designation_id = d.designation_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    for code, branch, dcode, exp in cur.fetchall():
        if code not in rows:
            continue
        rows[code]["dac_des"].setdefault(str(branch), {})[str(dcode)] = norm_date(exp)

    cur.execute(
        f"""
        SELECT ai.analysis_item_code, sub.short_name, sc.nd_107, sc.nd_107_expired_date
        FROM subcontractor_capability sc
        JOIN analysis_item ai ON sc.analysis_item_id = ai.analysis_item_id
        JOIN subcontractor sub ON sc.subcontractor_id = sub.subcontractor_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    for code, sub_name, nd107, exp in cur.fetchall():
        if code not in rows:
            continue
        rows[code]["sc"].append(
            {
                "subcontractor": sub_name,
                "nd_107": bool(nd107),
                "expired_date": norm_date(exp),
            }
        )

    cur.execute(
        f"""
        SELECT ai.analysis_item_code, sub.short_name, d.designation_code, scd.expired_date
        FROM subcontractor_capability_designation scd
        JOIN subcontractor_capability sc ON scd.subcontractor_capability_id = sc.subcontractor_capability_id
        JOIN analysis_item ai ON sc.analysis_item_id = ai.analysis_item_id
        JOIN subcontractor sub ON sc.subcontractor_id = sub.subcontractor_id
        JOIN designation d ON scd.designation_id = d.designation_id
        WHERE ai.analysis_item_code IN ({placeholders})
        """,
        codes,
    )
    for code, sub_name, dcode, exp in cur.fetchall():
        if code not in rows:
            continue
        key = f"{sub_name}|{dcode}"
        rows[code]["sc_des"][key] = norm_date(exp)

    return rows


def compare_text_field(
    issues: List[CheckIssue],
    code: str,
    field: str,
    excel_val: Any,
    db_val: Any,
    optional: bool = False,
) -> None:
    ex = norm_text(excel_val)
    db = norm_text(db_val)
    if optional and not ex:
        return
    if ex != db:
        issues.append(
            CheckIssue(code, "catalog", f"{field}: Excel={excel_val!r} DB={db_val!r}")
        )


def compare_item(excel: ExcelRow, db: Optional[Dict[str, Any]]) -> List[CheckIssue]:
    issues: List[CheckIssue] = []
    code = excel.code
    ef = excel.fields

    if db is None:
        return [CheckIssue(code, "catalog", "Không có analysis_item trên DB")]

    if norm_text(ef.get("name_vi")) != norm_text(db.get("name_vi")):
        issues.append(
            CheckIssue(code, "catalog", f"name_vi: Excel={ef.get('name_vi')!r} DB={db.get('name_vi')!r}")
        )

    if not tiptap_equal(ef.get("display_name_vi"), db.get("display_name_vi")):
        issues.append(
            CheckIssue(
                code,
                "display",
                "display_name_vi khác (Tiptap runs không khớp Excel)",
            )
        )

    for fld in (
        "name_en",
        "short_name",
        "analysis_group",
        "standard",
        "reference_method",
        "uom",
        "equipment",
        "sm_group",
        "sm",
        "lab_technique",
    ):
        compare_text_field(issues, code, fld, ef.get(fld), db.get(fld), optional=True)

    if ef.get("display_name_en") and not tiptap_equal(ef.get("display_name_en"), db.get("display_name_en")):
        issues.append(CheckIssue(code, "display", "display_name_en khác Tiptap"))
    if ef.get("display_short_name") and not tiptap_equal(ef.get("display_short_name"), db.get("display_short_name")):
        issues.append(CheckIssue(code, "display", "display_short_name khác Tiptap"))

    for tk in ("tat_normal", "tat_fast", "tat_urgent"):
        ex_t = (ef.get("tats") or {}).get(tk)
        db_t = (db.get("tats") or {}).get(tk)
        if ex_t != db_t:
            issues.append(CheckIssue(code, "tat", f"{tk}: Excel={ex_t} DB={db_t}"))

    ex_price = ef.get("unit_price")
    db_price = db.get("unit_price")
    if ex_price is not None and db_price is not None and ex_price != db_price:
        issues.append(CheckIssue(code, "catalog", f"unit_price: Excel={ex_price} DB={db_price}"))

    if excel.sheet == "Vietlabs":
        for branch, (ex_has, ex_date) in (ef.get("nd107") or {}).items():
            dac = (db.get("dac") or {}).get(branch)
            if ex_has:
                if not dac:
                    issues.append(CheckIssue(code, "capability", f"Thiếu DAC branch {branch} (Excel có NĐ107)"))
                elif not dac.get("nd_107"):
                    issues.append(CheckIssue(code, "capability", f"DAC {branch}: nd_107=false (Excel có ngày)"))
                elif dac.get("expired_date") != ex_date:
                    issues.append(
                        CheckIssue(
                            code,
                            "capability",
                            f"DAC {branch} ngày: Excel={ex_date} DB={dac.get('expired_date')}",
                        )
                    )
            elif dac and dac.get("nd_107"):
                issues.append(CheckIssue(code, "capability", f"DAC {branch} thừa (Excel Chưa có)"))

        for dcode, info in (ef.get("designations") or {}).items():
            if not info.get("has_value"):
                continue
            ex_date = info.get("expired_date")
            found = False
            for branch, des_map in (db.get("dac_des") or {}).items():
                if dcode in des_map:
                    found = True
                    if des_map[dcode] != ex_date:
                        issues.append(
                            CheckIssue(
                                code,
                                "designation",
                                f"{dcode}@{branch}: Excel={ex_date} DB={des_map[dcode]}",
                            )
                        )
            if not found:
                issues.append(CheckIssue(code, "designation", f"Thiếu designation {dcode} (Excel có ngày/x)"))

    if excel.sheet == "NTP":
        label = ef.get("ntp_label")
        sc_rows = db.get("sc") or []
        if label and not any(norm_text(s.get("subcontractor")) == norm_text(label) for s in sc_rows):
            issues.append(CheckIssue(code, "ntp", f"Thiếu subcontractor_capability cho NTP {label!r}"))

        ex_has, ex_date = ef.get("ntp_nd107") or (False, None)
        if ex_has and not sc_rows:
            issues.append(CheckIssue(code, "ntp", "Excel có NĐ107 NTP nhưng không có SC trên DB"))
        elif ex_has and sc_rows:
            match = next((s for s in sc_rows if s.get("expired_date") == ex_date), None)
            if match is None and sc_rows:
                issues.append(
                    CheckIssue(
                        code,
                        "ntp",
                        f"NĐ107 NTP ngày: Excel={ex_date} DB={[s.get('expired_date') for s in sc_rows]}",
                    )
                )

        for dcode, ex_date in (ef.get("ntp_designations") or {}).items():
            key_match = any(
                k.endswith(f"|{dcode}") and (db.get("sc_des") or {}).get(k) == ex_date
                for k in (db.get("sc_des") or {})
            )
            if not key_match:
                issues.append(
                    CheckIssue(code, "ntp", f"Thiếu/khác designation NTP {dcode}={ex_date!r}")
                )

    return issues


def random_check_capability_items(
    count: int,
    *,
    xlsx: Optional[str] = None,
    codes: Optional[List[str]] = None,
    seed: Optional[int] = None,
    connection_string: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Random spot-check N mã CT (hoặc danh sách codes cố định) — Excel v2 vs DB.

    Returns dict: selected_codes, results (per code), summary.
    """
    path = cwp.resolve_xlsx_arg(xlsx)
    if not path or not os.path.isfile(path):
        raise FileNotFoundError(f"Không tìm thấy workbook: {xlsx or path}")

    index = index_workbook(path)
    all_codes = sorted(index.keys())
    if not all_codes:
        raise RuntimeError("Workbook không có mã CT hợp lệ")

    if codes:
        selected = [c.strip() for c in codes if c.strip() in index]
        missing = [c.strip() for c in codes if c.strip() and c.strip() not in index]
    else:
        if count <= 0:
            raise ValueError("count phải > 0")
        rng = random.Random(seed)
        pool = list(all_codes)
        rng.shuffle(pool)
        selected = pool[: min(count, len(pool))]
        missing = []

    conn_str = connection_string or iai.CONNECTION_STRING
    conn = pyodbc.connect(conn_str)
    try:
        cur = conn.cursor()
        db_map = fetch_db_snapshots(cur, selected)
    finally:
        conn.close()

    per_code: Dict[str, Any] = {}
    all_issues: List[CheckIssue] = []
    for code in selected:
        excel_row = index[code]
        db_row = db_map.get(code)
        issues = compare_item(excel_row, db_row)
        all_issues.extend(issues)
        per_code[code] = {
            "sheet": excel_row.sheet,
            "row": excel_row.row_idx,
            "ok": len(issues) == 0,
            "issues": [{"area": i.area, "message": i.message} for i in issues],
        }

    return {
        "xlsx": path,
        "seed": seed,
        "requested": count if not codes else len(codes),
        "selected_codes": selected,
        "missing_codes": missing,
        "summary": {
            "checked": len(selected),
            "passed": sum(1 for v in per_code.values() if v["ok"]),
            "failed": sum(1 for v in per_code.values() if not v["ok"]),
            "issue_count": len(all_issues),
        },
        "results": per_code,
    }


def print_report(report: Dict[str, Any], verbose: bool = True) -> None:
    s = report["summary"]
    print("=" * 72)
    print("RANDOM CHECK NĂNG LỰC — Excel v2 vs DB")
    print("=" * 72)
    print(f"Workbook: {report['xlsx']}")
    if report.get("seed") is not None:
        print(f"Seed: {report['seed']}")
    print(f"Đã chọn: {len(report['selected_codes'])} mã")
    if report.get("missing_codes"):
        print(f"Mã không có trong Excel: {', '.join(report['missing_codes'])}")
    print(f"PASS: {s['passed']}/{s['checked']}  FAIL: {s['failed']}  Issues: {s['issue_count']}")
    print()

    for code in report["selected_codes"]:
        r = report["results"][code]
        status = "OK" if r["ok"] else "FAIL"
        print(f"[{status}] {code}  ({r['sheet']} row {r['row']})")
        if verbose and r["issues"]:
            for iss in r["issues"]:
                print(f"       [{iss['area']}] {iss['message']}")
    print()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Random spot-check mã CT: so sánh workbook v2 với DB"
    )
    parser.add_argument(
        "--count",
        "-n",
        type=int,
        default=None,
        help="Số mã CT ngẫu nhiên cần kiểm tra",
    )
    parser.add_argument(
        "--codes",
        default=None,
        help="Danh sách mã cố định, phân cách bởi dấu phẩy (ưu tiên hơn --count)",
    )
    parser.add_argument("--seed", type=int, default=None, help="Seed random (tái lập kết quả)")
    parser.add_argument("--xlsx", default=None, help="Workbook v2")
    parser.add_argument("--connection-string", default=None)
    parser.add_argument("--json-out", default=None, help="Ghi báo cáo JSON")
    parser.add_argument("--quiet", action="store_true", help="Chỉ in summary")
    args = parser.parse_args()

    codes = [c.strip() for c in args.codes.split(",")] if args.codes else None
    count = args.count
    if not codes:
        if count is None:
            try:
                raw = input("Nhập số lượng mã CT cần kiểm tra: ").strip()
                count = int(raw)
            except (EOFError, ValueError):
                print("Cần --count N hoặc --codes CT-xxx,...", file=sys.stderr)
                return 2
        if count <= 0:
            print("--count phải > 0", file=sys.stderr)
            return 2

    try:
        report = random_check_capability_items(
            count or 0,
            xlsx=args.xlsx,
            codes=codes,
            seed=args.seed,
            connection_string=args.connection_string,
        )
    except Exception as e:
        print(f"Lỗi: {e}", file=sys.stderr)
        return 1

    print_report(report, verbose=not args.quiet)

    if args.json_out:
        out_path = args.json_out
        os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"Đã ghi: {out_path}")

    return 0 if report["summary"]["failed"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
