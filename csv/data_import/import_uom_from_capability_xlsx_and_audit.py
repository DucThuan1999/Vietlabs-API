#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
(1) Đọc data/Capability.xlsx (sheet Vietlabs): cột Đơn vị tính (header "ĐVT" hoặc chứa "đơn vị tính")
    và thêm các giá trị từ cột **ĐVT khối lượng** (nếu có trên sheet).
    Thu thập giá trị ĐVT khác nhau (chuẩn hóa khớp import: import_analysis_item.normalize_text).
    INSERT vào dbo.unit_of_measure các ĐVT chưa có (khớp theo name_vi / name_en / mã — cùng logic augment_master_maps).
    Khớp phần audit "chuỗi Excel không có trong danh mục" (mục D/E).

(2) Sau khi import: quét so khớp Excel ↔ DB — chỉ tiêu (Mã CT-*) có ô ĐVT / ĐVT khối lượng
    nhưng analysis_item.unit_of_measure_id hoặc standard_quantity_unit_of_measure_id còn NULL.

  python3 import_uom_from_capability_xlsx_and_audit.py --dry-run
  python3 import_uom_from_capability_xlsx_and_audit.py
  python3 import_uom_from_capability_xlsx_and_audit.py --skip-import
  python3 import_uom_from_capability_xlsx_and_audit.py --skip-audit   # chỉ import ĐVT, không quét audit
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import import_analysis_item as iai
import import_analysis_items_capability_vietlabs_xlsx as cap

try:
    import pyodbc
except ImportError:
    pyodbc = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


_REPO_ROOT = Path(__file__).resolve().parents[3]
_DEFAULT_XLSX = _REPO_ROOT / "data" / "Capability.xlsx"

_RE_DVT = re.compile(r"^DVT-(\d+)$", re.IGNORECASE)


def _conn_str(args_conn: str) -> str:
    return (
        (args_conn or os.environ.get("VIETLABS_SQL_ODBC", "").strip() or iai.CONNECTION_STRING).strip()
    )


def _is_blank_uom_cell(s: str) -> bool:
    u = s.strip().upper().replace(" ", "")
    return u in ("", "NA", "N/A", "-", "--", "NONE", "NULL")


def fetch_table_columns(cur: "pyodbc.Cursor", table: str, schema: str = "dbo") -> Set[str]:
    cur.execute(
        """
        SELECT LOWER(COLUMN_NAME)
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """,
        schema,
        table,
    )
    return {r[0] for r in cur.fetchall()}


def load_uom_normalized_keys(cur: "pyodbc.Cursor") -> Set[str]:
    cur.execute(
        """
        SELECT name_vi, name_en, unit_of_measure_code
        FROM dbo.unit_of_measure
        """
    )
    keys: Set[str] = set()
    for nvi, nen, code in cur.fetchall():
        for x in (nvi, nen, code):
            if x is None or not str(x).strip():
                continue
            k = iai.normalize_text(str(x).strip())
            if k:
                keys.add(k)
    return keys


def max_dvt_numeric_suffix(cur: "pyodbc.Cursor") -> int:
    cur.execute(
        "SELECT unit_of_measure_code FROM dbo.unit_of_measure WHERE unit_of_measure_code LIKE N'DVT-%'"
    )
    mx = 0
    for (code,) in cur.fetchall():
        m = _RE_DVT.match((code or "").strip())
        if m:
            mx = max(mx, int(m.group(1), 10))
    return mx


def ordered_unique_uoms_from_column(
    rows: List[Tuple[Any, ...]],
    col_idx: int,
) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for row in rows:
        if col_idx >= len(row):
            continue
        raw = cap.cell_str(row[col_idx])
        if _is_blank_uom_cell(raw):
            continue
        nk = iai.normalize_text(raw)
        if not nk or nk in seen:
            continue
        seen.add(nk)
        out.append(raw.strip())
    return out


def merge_ordered_unique_uom_display(primary: List[str], extra: List[str]) -> List[str]:
    """Giữ thứ tự primary, nối thêm từ extra khi normalize_text chưa xuất hiện."""
    seen = {iai.normalize_text(x) for x in primary}
    out = list(primary)
    for s in extra:
        nk = iai.normalize_text(s)
        if not nk or nk in seen:
            continue
        seen.add(nk)
        out.append(s.strip())
    return out


def read_sheet_rows(xlsx_path: Path, sheet: str) -> Tuple[Tuple[Any, ...], List[Tuple[Any, ...]]]:
    if openpyxl is None:
        raise RuntimeError("Thiếu openpyxl. Cài: pip install openpyxl")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Không thấy sheet '{sheet}'. Có: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    it = ws.iter_rows(min_row=1, values_only=True)
    header = tuple(next(it) or ())
    data = list(it)
    return header, data


def build_ct_to_excel_uoms(
    header: Tuple[Any, ...],
    data: List[Tuple[Any, ...]],
) -> Tuple[Dict[str, Tuple[str, str]], List[str]]:
    """
    Returns:
      ct -> (main_uom_raw, std_qty_uom_raw) first non-empty wins per CT; warns on conflict.
      list of warning strings
    """
    cmap = cap.build_column_map(header)
    if "code" not in cmap:
        raise ValueError("Không tìm thấy cột Mã chỉ tiêu trên sheet.")
    ci = cmap["code"]
    ui = cmap.get("uom")
    if ui is None:
        raise ValueError("Không tìm thấy cột Đơn vị tính (ĐVT / đơn vị tính).")
    sqi = cmap.get("std_qty_uom")

    out: Dict[str, Tuple[str, str]] = {}
    warns: List[str] = []

    def merge_cell(prev: str, new_raw: str, label: str, ct: str) -> str:
        new_raw = new_raw.strip()
        if not new_raw or _is_blank_uom_cell(new_raw):
            return prev
        if not prev:
            return new_raw
        if iai.normalize_text(prev) != iai.normalize_text(new_raw):
            warns.append(
                f"{ct}: khác {label} giữa các dòng Excel ({prev!r} vs {new_raw!r}) — giữ giá trị đầu tiên."
            )
        return prev

    for row in data:
        if ci >= len(row):
            continue
        code = cap.cell_str(row[ci])
        if not cap.is_valid_ct_code(code):
            continue
        code = code.strip()
        u_main = ""
        u_sq = ""
        if ui < len(row):
            u_main = cap.cell_str(row[ui]).strip()
        if sqi is not None and sqi < len(row):
            u_sq = cap.cell_str(row[sqi]).strip()

        if code not in out:
            out[code] = ("", "")
        pm, ps = out[code]
        pm = merge_cell(pm, u_main, "ĐVT", code)
        ps = merge_cell(ps, u_sq, "ĐVT khối lượng", code)
        out[code] = (pm, ps)

    return out, warns


def fetch_analysis_item_uom_fk(
    cur: "pyodbc.Cursor", codes: List[str]
) -> Dict[str, Tuple[Optional[str], Optional[str]]]:
    """code -> (unit_of_measure_id, standard_quantity_unit_of_measure_id) as str GUID or None."""
    res: Dict[str, Tuple[Optional[str], Optional[str]]] = {}
    if not codes:
        return res
    CHUNK = 500
    for i in range(0, len(codes), CHUNK):
        chunk = codes[i : i + CHUNK]
        ph = ",".join(["?"] * len(chunk))
        cur.execute(
            f"""
            SELECT LTRIM(RTRIM(analysis_item_code)),
                   CONVERT(varchar(36), unit_of_measure_id),
                   CONVERT(varchar(36), standard_quantity_unit_of_measure_id)
            FROM dbo.analysis_item
            WHERE analysis_item_code IN ({ph})
            """,
            *chunk,
        )
        for c, uid, sid in cur.fetchall():
            res[str(c).strip()] = (
                str(uid).strip() if uid else None,
                str(sid).strip() if sid else None,
            )
    return res


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Import ĐVT từ Capability.xlsx + kiểm tra liên kết analysis_item ↔ ĐVT."
    )
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=_DEFAULT_XLSX,
        help="Đường dẫn Capability.xlsx (mặc định: <repo>/data/Capability.xlsx).",
    )
    ap.add_argument("--sheet", default="Vietlabs", help="Tên sheet (mặc định: Vietlabs).")
    ap.add_argument("--conn", default="", help="ODBC string hoặc dùng VIETLABS_SQL_ODBC / default trong import_analysis_item.")
    ap.add_argument("--dry-run", action="store_true", help="Không COMMIT INSERT.")
    ap.add_argument("--skip-import", action="store_true", help="Chỉ chạy bước kiểm tra liên kết.")
    ap.add_argument("--skip-audit", action="store_true", help="Không chạy bước quét sau import.")
    args = ap.parse_args()

    xlsx_path = args.xlsx.resolve()
    if not xlsx_path.is_file():
        print(f"Không tìm thấy file: {xlsx_path}", file=sys.stderr)
        return 2

    if pyodbc is None:
        print("Thiếu pyodbc.", file=sys.stderr)
        return 2

    conn_str = _conn_str(args.conn)
    if not conn_str:
        print("Thiếu connection string.", file=sys.stderr)
        return 2

    try:
        header, data = read_sheet_rows(xlsx_path, args.sheet)
    except Exception as ex:
        print(f"Lỗi đọc Excel: {ex}", file=sys.stderr)
        return 2

    cmap = cap.build_column_map(header)
    if "uom" not in cmap:
        print("Không tìm thấy cột Đơn vị tính (map uom).", file=sys.stderr)
        return 2

    uom_col = cmap["uom"]
    to_insert: List[str] = ordered_unique_uoms_from_column(data, uom_col)
    sq_col = cmap.get("std_qty_uom")
    if sq_col is not None:
        to_insert = merge_ordered_unique_uom_display(
            to_insert,
            ordered_unique_uoms_from_column(data, sq_col),
        )

    conn = pyodbc.connect(conn_str, autocommit=False)
    now = datetime.now(timezone.utc)
    try:
        cur = conn.cursor()
        uom_cols = fetch_table_columns(cur, "unit_of_measure")

        if not args.skip_import:
            db_keys = load_uom_normalized_keys(cur)
            mx = max_dvt_numeric_suffix(cur)
            pending: List[str] = []
            for disp in to_insert:
                nk = iai.normalize_text(disp)
                if nk in db_keys:
                    continue
                pending.append(disp)

            print("=" * 70)
            print("Bước 1 — Import đơn vị tính từ Capability.xlsx")
            print("=" * 70)
            print(f"File: {xlsx_path}  sheet={args.sheet!r}")
            print(
                f"Giá trị ĐVT khác nhau (cột chính"
                f"{' + cột ĐVT khối lượng' if sq_col is not None else ''}): {len(to_insert)}"
            )
            print(f"Cần INSERT (chưa có trên DB theo normalize_text): {len(pending)}")

            if pending:
                start = mx + 1
                print(f"MAX mã DVT-* hiện tại: DVT-{mx:03d}  ->  dòng mới từ DVT-{start:03d}")

            for i, disp in enumerate(pending):
                code = f"DVT-{start + i:03d}"
                nk = iai.normalize_text(disp)
                print(f"  {code}  name_vi/name_en={disp!r}")
                if args.dry_run:
                    continue
                uid = str(uuid.uuid4())
                if "unit_of_measure_id" not in uom_cols:
                    raise RuntimeError("Bảng unit_of_measure không có cột unit_of_measure_id.")
                cur.execute(
                    """
                    INSERT INTO dbo.unit_of_measure
                        (unit_of_measure_id, unit_of_measure_code, name_vi, name_en, status, created_at)
                    VALUES
                        (CAST(? AS UNIQUEIDENTIFIER), ?, ?, ?, N'Active', ?)
                    """,
                    (uid, code, disp, disp, now),
                )
                db_keys.add(nk)
                db_keys.add(iai.normalize_text(code))

            if args.dry_run:
                conn.rollback()
                print("\n[dry-run] Không ghi DB (bước import).")
            else:
                conn.commit()
                print(f"\nĐã commit: INSERT {len(pending)} đơn vị tính.")
        else:
            print("Bỏ qua bước import (--skip-import).")

        if args.skip_audit:
            return 0

        # --- Audit (đọc lại header/data đã có trong RAM) ---
        print("\n" + "=" * 70)
        print("Bước 2 — Kiểm tra liên kết chỉ tiêu (analysis_item) ↔ ĐVT")
        print("=" * 70)

        ct_map, warns = build_ct_to_excel_uoms(header, data)
        for w in warns[:50]:
            print(f"  [cảnh báo] {w}")
        if len(warns) > 50:
            print(f"  ... và {len(warns) - 50} cảnh báo khác.")

        codes = sorted(ct_map.keys())
        db_rows = fetch_analysis_item_uom_fk(cur, codes)

        missing_main: List[str] = []
        missing_sq: List[str] = []
        missing_row: List[str] = []

        for ct in codes:
            main_raw, sq_raw = ct_map[ct]
            row = db_rows.get(ct)
            if row is None:
                missing_row.append(ct)
                continue
            uid, sid = row
            if main_raw and not _is_blank_uom_cell(main_raw) and not uid:
                missing_main.append(ct)
            if sq_raw and not _is_blank_uom_cell(sq_raw) and not sid:
                missing_sq.append(ct)

        cur.execute(
            """
            SELECT COUNT(*) FROM dbo.analysis_item
            WHERE unit_of_measure_id IS NULL
            """
        )
        (cnt_null_uom,) = cur.fetchone()
        cur.execute(
            """
            SELECT COUNT(*) FROM dbo.analysis_item
            WHERE standard_quantity_unit_of_measure_id IS NULL
            """
        )
        (cnt_null_sq_uom,) = cur.fetchone()
        cur.execute("SELECT COUNT(*) FROM dbo.analysis_item")
        (cnt_ai,) = cur.fetchone()

        print(f"Tổng analysis_item: {cnt_ai}")
        print(f"  unit_of_measure_id IS NULL: {cnt_null_uom}")
        print(f"  standard_quantity_unit_of_measure_id IS NULL: {cnt_null_sq_uom}")
        print()
        print(f"Mã CT trên Excel (có ĐVT/ĐVT KL): {len(ct_map)}")
        print(f"  Không có dòng analysis_item trên DB: {len(missing_row)}")
        print(f"  Có ĐVT trên Excel nhưng unit_of_measure_id NULL: {len(missing_main)}")
        print(f"  Có ĐVT khối lượng trên Excel nhưng standard_quantity_unit_of_measure_id NULL: {len(missing_sq)}")

        def _show(label: str, items: List[str], limit: int = 40) -> None:
            if not items:
                return
            print(f"\n{label} (tối đa {limit} mã):")
            for c in items[:limit]:
                mm, sq = ct_map.get(c, ("", ""))
                dbu = db_rows.get(c)
                print(f"  {c}  excel_ĐVT={mm!r} excel_ĐVT_KL={sq!r}  DB_uom={dbu[0] if dbu else None} DB_sq_uom={dbu[1] if dbu else None}")
            if len(items) > limit:
                print(f"  ... +{len(items) - limit} mã nữa.")

        _show("CT có trên Excel nhưng chưa có analysis_item", missing_row)
        _show("Thiếu liên kết unit_of_measure_id", missing_main)
        _show("Thiếu liên kết standard_quantity_unit_of_measure_id", missing_sq)

        hint = (
            "\nGợi ý: chạy import analysis_item từ Capability "
            "(import_analysis_items_capability_vietlabs_xlsx.py) để gán FK; "
            "script này chỉ thêm master unit_of_measure."
        )
        if missing_main or missing_sq or missing_row:
            print(hint)

        return 0
    except Exception as ex:
        conn.rollback()
        print(f"Lỗi: {ex}", file=sys.stderr)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
