#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
So sánh "Phương pháp" từ data/Capability.xlsx (sheet Vietlabs) ↔ dbo.reference_method.

Mục tiêu:
- Liệt kê các phương pháp có trên XLSX nhưng chưa có trên DB
- Liệt kê các phương pháp có trên DB nhưng không thấy xuất hiện trên XLSX

Chuẩn hóa so khớp: dùng import_analysis_item.normalize_text (trim + upper + bỏ ký tự không phải word)
giống logic resolve_master/augment_master_maps trong import Capability.

  cd Vietlabs-API/csv/data_import
  python3 compare_reference_method_from_capability_xlsx.py
  python3 compare_reference_method_from_capability_xlsx.py --dry-run
  python3 compare_reference_method_from_capability_xlsx.py --xlsx /path/Capability.xlsx --sheet Vietlabs
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Set, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import import_analysis_item as iai
import import_analysis_items_capability_vietlabs_xlsx as cap_imp

try:
    import pyodbc
except ImportError:
    pyodbc = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


CONNECTION_STRING = (
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=192.168.2.21,49938;"
    "Database=VietLabs;"
    "UID=sa;"
    "PWD=Sbt@2024;"
    "TrustServerCertificate=yes;"
    "Encrypt=yes;"
    "Login Timeout=60;"
)

_REPO_ROOT = Path(__file__).resolve().parents[3]
_DEFAULT_XLSX = _REPO_ROOT / "data" / "Capability.xlsx"


def _conn_str(args_conn: str) -> str:
    return (args_conn or os.environ.get("VIETLABS_SQL_ODBC", "").strip() or CONNECTION_STRING).strip()


def _is_blank(s: str) -> bool:
    u = s.strip().upper().replace(" ", "")
    return u in ("", "NA", "N/A", "-", "--", "NONE", "NULL")


def cell_str(v: Any) -> str:
    return cap_imp.cell_str(v)


def read_sheet_rows(xlsx_path: Path, sheet: str) -> Tuple[Tuple[Any, ...], List[Tuple[Any, ...]]]:
    if openpyxl is None:
        raise RuntimeError("Thiếu openpyxl. Cài: pip install openpyxl")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Không thấy sheet '{sheet}'. Có: {', '.join(wb.sheetnames)}")
    ws = wb[sheet]
    it = ws.iter_rows(min_row=1, values_only=True)
    header = tuple(next(it) or ())
    data = list(it)
    return header, data


def ordered_unique_from_column(rows: List[Tuple[Any, ...]], col_idx: int) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for row in rows:
        if col_idx >= len(row):
            continue
        raw = cell_str(row[col_idx])
        if _is_blank(raw):
            continue
        nk = iai.normalize_text(raw)
        if not nk or nk in seen:
            continue
        seen.add(nk)
        out.append(raw.strip())
    return out


@dataclass(frozen=True)
class ReferenceMethodRow:
    reference_method_id: str
    code: str
    name_vi: str
    name_en: str


def _method_keys(r: ReferenceMethodRow) -> Set[str]:
    out: Set[str] = set()
    for x in (r.name_vi, r.name_en, r.code):
        if not x:
            continue
        k = iai.normalize_text(x)
        if k:
            out.add(k)
    return out


def load_db_reference_methods(
    cur: "pyodbc.Cursor",
) -> Tuple[List[ReferenceMethodRow], Set[str]]:
    """
    Returns:
      - list of reference_method rows
      - set of all normalized keys across rows (name_vi/name_en/code)
    """
    cur.execute(
        """
        SELECT
            CONVERT(varchar(36), reference_method_id) AS rid,
            LTRIM(RTRIM(ISNULL(reference_method_code, N''))) AS code,
            LTRIM(RTRIM(ISNULL(name_vi, N''))) AS name_vi,
            LTRIM(RTRIM(ISNULL(name_en, N''))) AS name_en
        FROM dbo.reference_method
        """
    )
    rows: List[ReferenceMethodRow] = []
    keys: Set[str] = set()
    for rid, code, nvi, nen in cur.fetchall():
        r = ReferenceMethodRow(
            reference_method_id=str(rid).strip(),
            code=str(code or "").strip(),
            name_vi=str(nvi or "").strip(),
            name_en=str(nen or "").strip(),
        )
        rows.append(r)
        keys |= _method_keys(r)
    return rows, keys


def fetch_analysis_item_reference_method_usage(
    cur: "pyodbc.Cursor",
    ids: List[str],
) -> Dict[str, int]:
    """reference_method_id -> count(analysis_item)."""
    out: Dict[str, int] = {}
    if not ids:
        return out
    CHUNK = 500
    for i in range(0, len(ids), CHUNK):
        chunk = ids[i : i + CHUNK]
        ph = ",".join(["CAST(? AS uniqueidentifier)"] * len(chunk))
        cur.execute(
            f"""
            SELECT CONVERT(varchar(36), reference_method_id) AS rid,
                   COUNT(*) AS cnt
            FROM dbo.analysis_item
            WHERE reference_method_id IN ({ph})
            GROUP BY reference_method_id
            """,
            *chunk,
        )
        for rid, cnt in cur.fetchall():
            out[str(rid).strip()] = int(cnt or 0)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="So sánh Phương pháp: Capability.xlsx ↔ DB reference_method")
    ap.add_argument("--xlsx", type=Path, default=_DEFAULT_XLSX)
    ap.add_argument("--sheet", default="Vietlabs")
    ap.add_argument("--conn", default="")
    ap.add_argument("--max-print", type=int, default=120)
    args = ap.parse_args()

    xlsx_path = args.xlsx.resolve()
    if not xlsx_path.is_file():
        print(f"Không tìm thấy: {xlsx_path}", file=sys.stderr)
        return 2

    conn_str = _conn_str(args.conn)
    if not conn_str or pyodbc is None:
        print("Thiếu connection string hoặc pyodbc.", file=sys.stderr)
        return 2

    try:
        header, data = read_sheet_rows(xlsx_path, args.sheet)
    except Exception as ex:
        print(f"Lỗi đọc Excel: {ex}", file=sys.stderr)
        return 2

    colmap = cap_imp.build_column_map(header)
    if "reference_method" not in colmap:
        print("Không tìm thấy cột 'Phương pháp' trên XLSX.", file=sys.stderr)
        return 2

    rm_col = colmap["reference_method"]
    excel_list = ordered_unique_from_column(data, rm_col)
    excel_keys = {iai.normalize_text(x) for x in excel_list if iai.normalize_text(x)}
    excel_disp = {iai.normalize_text(x): x for x in excel_list if iai.normalize_text(x)}

    conn = pyodbc.connect(conn_str)
    try:
        cur = conn.cursor()
        db_rows, db_keys = load_db_reference_methods(cur)
    finally:
        conn.close()

    only_xlsx = sorted(excel_keys - db_keys)
    db_only_methods: List[ReferenceMethodRow] = []
    for r in db_rows:
        if _method_keys(r).isdisjoint(excel_keys):
            db_only_methods.append(r)

    lim = max(0, int(args.max_print))

    print(f"XLSX: {xlsx_path}  sheet={args.sheet!r}")
    print(f"Phương pháp duy nhất (XLSX): {len(excel_keys)}")
    print(f"Key duy nhất (DB: name_vi/name_en/code): {len(db_keys)}")
    print()

    print(f"=== Có trên XLSX, chưa có trên DB ({len(only_xlsx)}) ===")
    if not only_xlsx:
        print("(không có)")
    else:
        show = only_xlsx if lim == 0 else only_xlsx[:lim]
        for k in show:
            print(excel_disp.get(k, k))
        if lim > 0 and len(only_xlsx) > lim:
            print(f"... +{len(only_xlsx) - lim} dòng nữa")

    print()
    print(f"=== Có trên DB, không thấy trên XLSX ({len(db_only_methods)}) ===")
    if not db_only_methods:
        print("(không có)")
    else:
        conn = pyodbc.connect(conn_str)
        try:
            cur = conn.cursor()
            usage = fetch_analysis_item_reference_method_usage(
                cur, [r.reference_method_id for r in db_only_methods]
            )
        finally:
            conn.close()

        used = [r for r in db_only_methods if usage.get(r.reference_method_id, 0) > 0]
        unused = [r for r in db_only_methods if usage.get(r.reference_method_id, 0) == 0]

        print(f"- Đang được dùng bởi analysis_item: {len(used)}")
        print(f"- Không FK tới analysis_item:      {len(unused)}")

        show = db_only_methods if lim == 0 else db_only_methods[:lim]
        for r in show:
            cnt = usage.get(r.reference_method_id, 0)
            label = r.name_vi or r.name_en or ""
            pref = f"[{r.code}] " if r.code else ""
            print(f"  {pref}{label}  (analysis_item FK={cnt})")
        if lim > 0 and len(db_only_methods) > lim:
            print(f"... +{len(db_only_methods) - lim} dòng nữa")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

